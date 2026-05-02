"""Tabular invoice (XLSX) parser — Jalon 5.3.5.

Some lessors export invoice data as a flat XLSX listing one row per
``(plate × invoice × line item)``. Ayvens calls it « Etat des dépenses »,
Arval/autre loueur may have similar exports under different names. The
shape is consistent enough that we can parse them all the same way :

- one column for the plate (``Immatriculation`` / ``Immat`` / …)
- one column for the contract number (``N° contrat`` / ``Contrat``)
- one column for the invoice date (``Date Facture`` / ``Date facture``)
- one column for the rubrique label (``Libellé prestation`` / ``Poste``)
- one column for the HT amount (``Montant HT``)
- one column for the TTC amount (``Montant TTC``)

This parser reads such files and produces the SAME DataFrame shape as
:func:`pdf_parser.parse_factures_to_dataframe` so the contract engine
treats them interchangeably — the file lands under one of the existing
``*_facture_pdf`` slugs and all the YAML rules already declared for
those slugs apply unchanged. No new YAML to write.

How rubrique → Revio field mapping works
----------------------------------------
We reuse ``pdf_parser.classify_rubriques`` (same whitelist/blacklist
regex over ``rubriques_facture.yml``). One XLSX line = one ``Rubrique``
record ; we group all lines for a given ``(plate, number)``, classify
their labels, sum the amounts per ``revio_field`` (with ``_ht`` /
``_ttc`` flavours), and return one row per contract — same as the PDF
parser's per-block aggregation.

Multi-month dedup
-----------------
A real « Etat des dépenses » contains 12+ months of data. To match
:func:`parse_factures_to_dataframe` semantics (R4 — keep most recent
facture per contract), we filter rows to the LATEST ``Date Facture``
per ``(plate, number)`` group before aggregating. That way the engine
sees the same shape as if a single (latest) PDF facture was uploaded.

Public API
----------
* :func:`parse_etat_depenses_to_dataframe` — main entry, mirrors
  ``parse_factures_to_dataframe`` signature.
* :func:`is_etat_depenses_shape` — column fingerprint detector for the
  upload pipeline.
* :func:`detect_lessor_from_filename` — best-effort lessor inference.
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Optional

import openpyxl
import pandas as pd

from .pdf_parser import (
    PRICE_FIELDS,
    Rubrique,
    classify_rubriques,
)


# =============================================================================
# Column fingerprint
# =============================================================================
#
# Each entry = (canonical_field, alias_set). The aliases are matched
# case-insensitive after stripping accents. The list is intentionally
# small ; if a future lessor renames a column we add the alias here.

# IMPORTANT — these are LISTS (ordered), not sets. The order matters
# because :func:`_resolve_columns` picks the FIRST alias whose label is
# present in the file's header. If you put a generic alias before a
# specific one, the generic alias will hijack the match. Concrete
# example : Ayvens « Etat des dépenses » has BOTH ``Libellé produit``
# (vehicle description) and ``Libellé prestation`` (actual rubrique).
# We want ``Libellé prestation`` to win, so it's listed first.
_FIELD_ALIASES: dict[str, list[str]] = {
    "plate": [
        "immatriculation", "n° immat", "n immat", "no immat", "immat",
        "plaque",
    ],
    "number": [
        "n° contrat", "n contrat", "no contrat", "numero contrat",
        "ref contrat", "contrat",
    ],
    "facture_date": [
        "date facture", "date de facture", "date_facture",
    ],
    "facture_number": [
        "n° facture", "n facture", "no facture",
    ],
    "rubrique_label": [
        # MOST-SPECIFIC FIRST — see module-level note above.
        "libellé prestation", "libelle prestation",
        "poste de depense", "poste dépense", "poste depense",
        "rubrique", "désignation", "designation",
        # Fallbacks (more generic — last resort) :
        "libellé produit", "libelle produit",
    ],
    "montant_ht": [
        "montant ht", "montant_ht", "ht",
    ],
    "montant_ttc": [
        "montant ttc", "montant_ttc", "ttc",
    ],
}

# Minimal subset of fields that MUST be present for the file to be
# recognised as an « Etat des dépenses » export. We don't require facture
# number / TTC because some exports lack them.
_REQUIRED_FIELDS = ("plate", "number", "facture_date", "rubrique_label", "montant_ht")


def _strip_accents(s: str) -> str:
    import unicodedata
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if not unicodedata.combining(c))


def _norm_col(s: Any) -> str:
    """Lower + strip accents + collapse spaces. Used for header matching."""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", _strip_accents(str(s)).strip().lower())


def _resolve_columns(headers: list[str]) -> dict[str, Optional[str]]:
    """Map our canonical field names to the actual header label in the file.

    Returns ``{canonical_field: header_label_or_None}``. Headers that
    don't match any alias are simply ignored.
    """
    norm_to_orig: dict[str, str] = {}
    for h in headers:
        if h is None:
            continue
        norm_to_orig.setdefault(_norm_col(h), str(h))
    out: dict[str, Optional[str]] = {}
    for canonical, aliases in _FIELD_ALIASES.items():
        out[canonical] = None
        for alias in aliases:
            if alias in norm_to_orig:
                out[canonical] = norm_to_orig[alias]
                break
    return out


def is_etat_depenses_shape(headers: Iterable[str]) -> bool:
    """True iff ``headers`` look like an « Etat des dépenses » export.

    We require the minimal column set defined in ``_REQUIRED_FIELDS``.
    Used by the upload pipeline to auto-route the file.
    """
    resolved = _resolve_columns(list(headers))
    return all(resolved.get(f) for f in _REQUIRED_FIELDS)


# =============================================================================
# Header row detection
# =============================================================================
#
# Real exports start with a few title / metadata rows before the actual
# header. We scan the first 15 rows looking for one whose contents pass
# :func:`is_etat_depenses_shape`.

_HEADER_SCAN_LIMIT = 15


def _find_header_row(rows: list[tuple]) -> Optional[int]:
    """Return the 0-based index of the header row in ``rows``, or None."""
    for i, row in enumerate(rows[:_HEADER_SCAN_LIMIT]):
        cells = [str(c) for c in row if c is not None]
        if not cells:
            continue
        if is_etat_depenses_shape(cells):
            return i
    return None


# =============================================================================
# Lessor detection
# =============================================================================

_LESSOR_FILENAME_HINTS = {
    "ayvens": "ayvens",
    "ald":    "ayvens",  # ALD = ancien nom Ayvens
    "leaseplan": "ayvens",
    "arval":  "arval",
}


def detect_lessor_from_filename(filename: str) -> str:
    """Return ``ayvens`` / ``arval`` / ``autre`` based on filename hints."""
    low = (filename or "").lower()
    for key, lessor in _LESSOR_FILENAME_HINTS.items():
        if key in low:
            return lessor
    return "autre"


def lessor_to_slug(lessor: str) -> str:
    """Map the lessor inferred from the file to its engine source slug.

    The slug is one of the ``*_facture_pdf`` slugs the contract engine
    already knows — that way no new YAML rules are needed.
    """
    table = {
        "ayvens": "ayvens_facture_pdf",
        "arval":  "arval_facture_pdf",
        "autre":  "autre_loueur_facture_pdf",
    }
    return table.get(lessor, "autre_loueur_facture_pdf")


# =============================================================================
# Date / amount parsing
# =============================================================================

def _parse_date(v: Any) -> Optional[datetime]:
    """Parse a Date Facture cell to a datetime. Returns None on failure."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d",
        "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y",
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


_NUMBER_CLEAN_RE = re.compile(r"[^0-9,.\-]")


def _parse_amount(v: Any) -> Optional[float]:
    """Parse a montant cell. Handles fr (1 234,56) and en (1,234.56) formats."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = _NUMBER_CLEAN_RE.sub("", str(v).strip())
    if not s:
        return None
    # If both ',' and '.' present: assume the last one is the decimal sep.
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        # Pure FR format
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


# =============================================================================
# Reading
# =============================================================================

def _read_xlsx(path: str | Path) -> tuple[list[str], list[list[Any]]]:
    """Read sheet 0 and return ``(header_row, data_rows)``.

    Raises :class:`ValueError` when the file has no recognisable header.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = _find_header_row(rows)
    if header_idx is None:
        raise ValueError(
            f"Aucun en-tête « Etat des dépenses » trouvé dans {path}"
        )
    header = [c for c in rows[header_idx]]
    data = [list(r) for r in rows[header_idx + 1:] if any(c is not None for c in r)]
    return header, data


# =============================================================================
# Aggregation — group by (plate, number), latest date wins, sum rubriques
# =============================================================================

def _normalise_plate(p: Any) -> Optional[str]:
    """Normalised plate display string (with hyphens preserved if any)."""
    if p is None:
        return None
    s = str(p).strip()
    return s or None


def _normalise_number(n: Any) -> Optional[str]:
    if n is None:
        return None
    s = str(n).strip()
    return s or None


def _aggregate_rubriques(
    plate: str,
    number: Optional[str],
    rubriques: list[Rubrique],
    facture_date: Optional[datetime],
    use_ttc: bool,
) -> dict[str, Any]:
    """Mirror of ``pdf_parser.block_to_row`` for one (plate, number) group.

    Sums the classified rubriques into a flat output row with both
    ``_ht`` and ``_ttc`` flavours plus the engine-facing column.
    """
    row: dict[str, Any] = {
        "plate": plate,
        "number": number,
        "lessor": None,             # filled by caller
        "facture_date": facture_date,
        "durationMonths": None,     # not available in flat XLSX exports
        "maxMileage": None,
        "restitutionDate": None,
        "startDate": None,
        "endDate": None,
    }
    for f in PRICE_FIELDS:
        row[f + "_ht"] = None
        row[f + "_ttc"] = None

    total_ht = 0.0
    total_ttc = 0.0
    had_total_contribution = False
    unknown: list[str] = []

    for r in rubriques:
        if r.blacklisted:
            continue
        if r.unknown:
            unknown.append(r.label)
            continue
        if r.ht is not None:
            total_ht += r.ht
        if r.ttc is not None:
            total_ttc += r.ttc
        had_total_contribution = True
        if r.revio_field:
            key_ht = r.revio_field + "_ht"
            key_ttc = r.revio_field + "_ttc"
            row[key_ht] = (row.get(key_ht) or 0.0) + (r.ht or 0.0)
            row[key_ttc] = (row.get(key_ttc) or 0.0) + (r.ttc or 0.0)

    row["totalPrice_ht"] = round(total_ht, 2) if had_total_contribution else None
    row["totalPrice_ttc"] = round(total_ttc, 2) if had_total_contribution else None
    row["_unknown_rubriques"] = unknown

    if use_ttc:
        for f in PRICE_FIELDS + ["totalPrice"]:
            row[f] = row.get(f + "_ttc")
    else:
        for f in PRICE_FIELDS + ["totalPrice"]:
            row[f] = row.get(f + "_ht")
    return row


# =============================================================================
# Public entry point
# =============================================================================

def parse_etat_depenses_to_dataframe(
    xlsx_paths: Iterable[str | Path],
    whitelist: list[dict],
    blacklist: list[dict],
    lessor_hint: Optional[str] = None,
    assume_ttc: bool = True,
) -> pd.DataFrame:
    """Parse one or more « Etat des dépenses » XLSX into one DataFrame.

    The output has the SAME columns as
    :func:`pdf_parser.parse_factures_to_dataframe` so the contract
    engine consumes both interchangeably.

    Multi-file behaviour
    --------------------
    ``xlsx_paths`` may contain several files. They are concatenated and
    deduplicated by ``(plate, number)`` keeping the most recent
    ``facture_date`` — same R4 dedup rule as the PDF parser.

    ``lessor_hint`` overrides the per-file lessor inference. When None
    the lessor is detected from the filename (Ayvens / Arval / autre).
    """
    all_rows: list[dict[str, Any]] = []

    for path in xlsx_paths:
        try:
            header, data = _read_xlsx(path)
        except (ValueError, OSError) as e:
            # Caller is responsible for surfacing this; we skip the file.
            print(f"[invoice_xlsx_parser] skip {path}: {e}")
            continue

        col_map = _resolve_columns(header)
        # Index helpers — `_find_idx` is None-safe.
        def _find_idx(canonical: str) -> Optional[int]:
            label = col_map.get(canonical)
            if not label:
                return None
            try:
                return header.index(label)
            except ValueError:
                return None

        i_plate   = _find_idx("plate")
        i_number  = _find_idx("number")
        i_date    = _find_idx("facture_date")
        i_label   = _find_idx("rubrique_label")
        i_ht      = _find_idx("montant_ht")
        i_ttc     = _find_idx("montant_ttc")

        # Refuse files that don't have the minimum set; we already
        # checked the header row but a real-world file might have
        # blank/duplicate headers that confused us.
        if i_plate is None or i_number is None or i_label is None or i_ht is None:
            print(f"[invoice_xlsx_parser] missing required columns in {path}")
            continue

        lessor = lessor_hint or detect_lessor_from_filename(str(path))

        # Step 1 : per (plate, number) group keep the latest facture_date.
        latest_date_by_key: dict[tuple[str, str], datetime] = {}
        for row in data:
            plate = _normalise_plate(row[i_plate]) if i_plate < len(row) else None
            number = _normalise_number(row[i_number]) if i_number < len(row) else None
            if not plate:
                continue
            key = (plate, number or "")
            d = _parse_date(row[i_date]) if i_date is not None and i_date < len(row) else None
            if d is None:
                continue
            cur = latest_date_by_key.get(key)
            if cur is None or d > cur:
                latest_date_by_key[key] = d

        # Step 2 : collect rubriques for the latest date per group.
        rubriques_by_key: dict[tuple[str, str], list[Rubrique]] = {}
        for row in data:
            plate = _normalise_plate(row[i_plate]) if i_plate < len(row) else None
            number = _normalise_number(row[i_number]) if i_number < len(row) else None
            if not plate:
                continue
            key = (plate, number or "")
            d = _parse_date(row[i_date]) if i_date is not None and i_date < len(row) else None
            target_date = latest_date_by_key.get(key)
            # When a key has no parseable date at all, target_date is None and
            # we still aggregate everything (rare but keeps small test files
            # working).
            if target_date is not None and d != target_date:
                continue

            label = row[i_label] if i_label < len(row) else None
            if not label or not str(label).strip():
                continue
            ht  = _parse_amount(row[i_ht])  if i_ht  < len(row) else None
            ttc = _parse_amount(row[i_ttc]) if i_ttc is not None and i_ttc < len(row) else None
            rubriques_by_key.setdefault(key, []).append(
                Rubrique(label=str(label).strip(), ht=ht, ttc=ttc)
            )

        # Step 3 : classify + aggregate per group → one row per contract.
        for (plate, number), rubriques in rubriques_by_key.items():
            classify_rubriques(rubriques, whitelist, blacklist)
            facture_date = latest_date_by_key.get((plate, number))
            row = _aggregate_rubriques(
                plate=plate,
                number=number or None,
                rubriques=rubriques,
                facture_date=facture_date,
                use_ttc=assume_ttc,
            )
            row["lessor"] = lessor
            row["_source_xlsx"] = str(path)
            all_rows.append(row)

    if not all_rows:
        return pd.DataFrame()

    df = pd.DataFrame(all_rows)
    # Cross-file dedup : same (plate, number) might be present in multiple
    # XLSX exports — keep the most recent facture_date row.
    df = df.sort_values(
        by=["plate", "number", "facture_date"],
        ascending=[True, True, False],
        na_position="last",
    )
    df = df.drop_duplicates(subset=["plate", "number"], keep="first")
    df = df.reset_index(drop=True)
    return df
