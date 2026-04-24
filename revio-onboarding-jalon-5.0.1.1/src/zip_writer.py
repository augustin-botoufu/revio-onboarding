"""Package the engine output into a single downloadable zip (Jalon 3.0 + 4.2).

The UI requirement is « un seul bouton, un seul zip » with the following
layout::

    revio_import_<client>_<ts>/
    ├── onboarding_complet.xlsx      (master workbook, multi-tabs V + C)
    ├── rapport.xlsx                  (per-cell source / anomalies véhicules)
    ├── contracts_errors.xlsx         (anomalies contrats — si dispo)
    ├── vehicles/
    │   ├── vehicles_tous.csv         (every row, Revio header order)
    │   ├── vehicles_<slug>.csv       (one file per fleet)
    │   └── vehicles_non-rattache.csv (plates with no fleet, if any)
    ├── contracts/
    │   ├── contracts_tous.csv
    │   ├── contracts_<slug>.csv      (per fleet, plate-based)
    │   └── contracts_orphelins.csv   (contrats sans match client_file)
    └── _lineage/                     (optional — parquet/jsonl sidecars)
        ├── vehicle.parquet
        └── contract.parquet

If no fleet mapping is active we skip the per-fleet tabs and CSVs, and
ship only a single ``*_tous.csv`` plus the « tous » tab in the master
Excel. No empty artifacts.

Jalon 4.2 adds contract outputs + arbitrary ``extra_files`` (used for
lineage sidecars shared with the Jalon 5.0 LLM assistant).
"""

from __future__ import annotations

import io
import re
import zipfile
from datetime import datetime
from typing import Iterable, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .fleet_segmentation import (
    FleetMapping,
    UNASSIGNED_FLEET,
    safe_sheet_name,
    slugify_fleet_name,
    split_df_by_fleet,
)
from .normalizers import plate_for_matching
from .schemas import header_for


# =============================================================================
# Revio CSV (per-schema, per-fleet)
# =============================================================================

def revio_csv_bytes(df: pd.DataFrame, schema_name: str) -> bytes:
    """Serialize ``df`` to CSV in the exact Revio header order.

    The Vehicle schema has an empty-named placeholder column between
    ``co2gKm`` and ``registrationIssueCountryCode`` (mirror of the NAT01
    template). We preserve it verbatim in the output header.

    Encodes as UTF-8 with BOM (``utf-8-sig``) so Excel opens it in the
    correct encoding when double-clicked.
    """
    header = header_for(schema_name)
    out = pd.DataFrame(index=range(len(df)))
    empty_counter = 0
    for col in header:
        if col == "":
            empty_counter += 1
            out[f"__empty_{empty_counter}"] = ""
        else:
            if col in df.columns:
                out[col] = df[col].values
            else:
                out[col] = ""
    buf = io.StringIO()
    out.to_csv(buf, index=False)
    lines = buf.getvalue().splitlines()
    # Restore the true header (with any empty slots).
    lines[0] = ",".join(header)
    return ("\n".join(lines) + "\n").encode("utf-8-sig")


# =============================================================================
# Master Excel workbook (multi-tab)
# =============================================================================

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", start_color="1F2937")  # slate-800
_TOC_TITLE_FONT = Font(bold=True, size=14)


def _autosize(ws, max_width: int = 48) -> None:
    """Rough column autosizing — caps at ``max_width`` chars."""
    for col_idx, col_cells in enumerate(ws.iter_cols(values_only=False), start=1):
        best = 10
        for cell in col_cells:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            if length > best:
                best = length
        ws.column_dimensions[get_column_letter(col_idx)].width = min(best + 2, max_width)


def _write_df_to_sheet(ws, df: pd.DataFrame, schema_name: str) -> None:
    """Dump ``df`` into sheet ``ws`` respecting the Revio header order."""
    header = [c if c != "" else " " for c in header_for(schema_name)]
    ws.append(header)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for _, row in df.iterrows():
        out_row: list = []
        for name in header_for(schema_name):
            if name == "":
                out_row.append("")
                continue
            v = row.get(name, "")
            # pandas NaN → empty string for display
            if v is None or (isinstance(v, float) and pd.isna(v)):
                out_row.append("")
            else:
                out_row.append(v)
        ws.append(out_row)

    ws.freeze_panes = "A2"
    _autosize(ws)


def _build_toc(
    wb: Workbook,
    plan: list[tuple[str, str, int]],  # [(sheet_name, description, n_rows)]
    client_name: str,
) -> None:
    """Insert a « Sommaire » sheet at the top of the workbook."""
    ws = wb.create_sheet("Sommaire", 0)
    ws["A1"] = f"Onboarding Revio — {client_name}"
    ws["A1"].font = _TOC_TITLE_FONT
    ws["A2"] = f"Généré le {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color="6B7280")
    ws.append([])
    ws.append(["Onglet", "Contenu", "Lignes"])
    for cell in ws[4]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for sheet_name, desc, n_rows in plan:
        ws.append([sheet_name, desc, n_rows])
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 10
    ws.freeze_panes = "A5"


def build_master_xlsx(
    *,
    client_name: str,
    vehicle_df: Optional[pd.DataFrame],
    vehicle_fleet_mapping: Optional[FleetMapping],
    contract_df: Optional[pd.DataFrame] = None,
    contract_orphan_df: Optional[pd.DataFrame] = None,
    contract_fleet_mapping: Optional[FleetMapping] = None,
) -> bytes:
    """Build a workbook with one tab per (schema, fleet) + a global tab.

    Jalon 4.2: populates both ``vehicle`` and ``contract`` schemas. Each
    contributes:
        - 1 « tous » tab
        - N tabs (one per fleet) if a fleet mapping is active
    For contracts, an additional « orphelins » tab lists rows present in
    loueur sources but missing from client_file (R6 cross-check).
    """
    wb = Workbook()
    # Remove the default sheet; we'll rebuild a clean ToC at index 0.
    wb.remove(wb.active)

    plan: list[tuple[str, str, int]] = []

    # ---- VEHICLE ----
    if vehicle_df is not None and not vehicle_df.empty:
        ws_all = wb.create_sheet(safe_sheet_name("vehicles — tous"))
        _write_df_to_sheet(ws_all, vehicle_df, "vehicle")
        plan.append((ws_all.title, "Tous les véhicules", len(vehicle_df)))

        buckets = split_df_by_fleet(vehicle_df, vehicle_fleet_mapping)
        if vehicle_fleet_mapping is not None and vehicle_fleet_mapping.is_active:
            for fleet_name in _fleet_iter_order(buckets.keys()):
                sub = buckets[fleet_name]
                if sub.empty:
                    continue
                sheet_name = safe_sheet_name(f"vehicles — {fleet_name}")
                ws = wb.create_sheet(sheet_name)
                _write_df_to_sheet(ws, sub, "vehicle")
                desc = f"Véhicules flotte « {fleet_name} »"
                plan.append((sheet_name, desc, len(sub)))

    # ---- CONTRACT ----
    if contract_df is not None and not contract_df.empty:
        ws_all = wb.create_sheet(safe_sheet_name("contrats — tous"))
        _write_df_to_sheet(ws_all, contract_df, "contract")
        plan.append((ws_all.title, "Tous les contrats", len(contract_df)))

        # Per-fleet tabs (plate-based mapping shared with Vehicle).
        if contract_fleet_mapping is not None and contract_fleet_mapping.is_active:
            c_buckets = _split_contracts_by_fleet(contract_df, contract_fleet_mapping)
            for fleet_name in _fleet_iter_order(c_buckets.keys()):
                sub = c_buckets[fleet_name]
                if sub.empty:
                    continue
                sheet_name = safe_sheet_name(f"contrats — {fleet_name}")
                ws = wb.create_sheet(sheet_name)
                _write_df_to_sheet(ws, sub, "contract")
                desc = f"Contrats flotte « {fleet_name} »"
                plan.append((sheet_name, desc, len(sub)))

    if contract_orphan_df is not None and not contract_orphan_df.empty:
        sheet_name = safe_sheet_name("contrats — orphelins")
        ws = wb.create_sheet(sheet_name)
        # Orphans keep their raw loueur columns (no schema re-ordering).
        cols = list(contract_orphan_df.columns)
        ws.append(cols)
        for cell in ws[1]:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for _, row in contract_orphan_df.iterrows():
            ws.append([
                "" if row.get(c) is None or (isinstance(row.get(c), float) and pd.isna(row.get(c)))
                else row.get(c)
                for c in cols
            ])
        ws.freeze_panes = "A2"
        _autosize(ws)
        plan.append((sheet_name, "Contrats présents loueur mais absents client_file", len(contract_orphan_df)))

    if not plan:
        # Nothing to export — create a single info sheet so the xlsx is valid.
        ws = wb.create_sheet("Vide")
        ws["A1"] = "Aucune donnée à exporter."
        plan.append(("Vide", "Aucune donnée", 0))

    _build_toc(wb, plan, client_name=client_name)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _split_contracts_by_fleet(
    contract_df: pd.DataFrame,
    fleet_mapping: FleetMapping,
) -> dict[str, pd.DataFrame]:
    """Split contract rows by fleet using the plate column.

    Reuses ``fleet_mapping.plate_to_fleet`` (built from the Vehicle side):
    a contract inherits the fleet of its plate. Contracts whose plate is
    missing from the mapping fall into the UNASSIGNED_FLEET bucket.

    Works whether ``contract_df`` is indexed by composite key ``plate|number``
    or by a default integer index — we resolve the plate from the ``plate``
    column (falling back to the index parse if absent).
    """
    if contract_df is None or contract_df.empty:
        return {}

    # Resolve the plate per-row: prefer the explicit column, otherwise parse
    # the composite-key index ("plate|number" → "plate").
    if "plate" in contract_df.columns:
        plate_series = contract_df["plate"]
    elif "registrationPlate" in contract_df.columns:
        plate_series = contract_df["registrationPlate"]
    else:
        plate_series = pd.Series(
            [str(idx).split("|", 1)[0] if idx is not None else ""
             for idx in contract_df.index],
            index=contract_df.index,
        )

    # Normalize via the shared helper so lookups agree with the Vehicle side.
    plates_norm = plate_series.map(lambda p: plate_for_matching(p) or "")
    p2f = fleet_mapping.plate_to_fleet

    assigned = plates_norm.map(lambda p: p2f.get(p, UNASSIGNED_FLEET) if p else UNASSIGNED_FLEET)
    out: dict[str, pd.DataFrame] = {}
    for name, sub in contract_df.groupby(assigned, sort=False):
        out[str(name)] = sub.copy()
    return out


# =============================================================================
# Top-level zip builder
# =============================================================================

def _fleet_iter_order(names: Iterable[str]) -> list[str]:
    """Iterate fleets in a stable display order: real names alpha, then
    the « non rattaché » bucket last. Skips empty-string bucket."""
    real = sorted(n for n in names if n and n != UNASSIGNED_FLEET)
    out = list(real)
    if UNASSIGNED_FLEET in names:
        out.append(UNASSIGNED_FLEET)
    return out


def _safe_client_slug(client_name: str) -> str:
    """Filename-safe version of the client name."""
    s = slugify_fleet_name(client_name) if client_name else ""
    return s or "client"


def build_output_zip(
    *,
    client_name: str,
    vehicle_df: Optional[pd.DataFrame] = None,
    vehicle_fleet_mapping: Optional[FleetMapping] = None,
    report_xlsx_bytes: Optional[bytes] = None,
    contract_df: Optional[pd.DataFrame] = None,
    contract_orphan_df: Optional[pd.DataFrame] = None,
    contract_errors_xlsx_bytes: Optional[bytes] = None,
    contract_fleet_mapping: Optional[FleetMapping] = None,
    extra_files: Optional[dict[str, bytes]] = None,
    timestamp: Optional[datetime] = None,
) -> tuple[bytes, str]:
    """Assemble the final zip and return ``(bytes, filename)``.

    The filename embeds the client slug and a UTC-ish timestamp so two
    successive downloads don't collide in the user's Downloads folder.

    Jalon 4.2 additions:
    - ``contract_df`` / ``contract_orphan_df`` → ``contracts/*.csv`` and
      contract tabs in the master workbook.
    - ``contract_errors_xlsx_bytes`` → ``contracts_errors.xlsx`` at the root.
    - ``contract_fleet_mapping`` → per-fleet contract CSVs (plate-based).
    - ``extra_files`` → ``{relative_path: bytes}`` written verbatim inside
      the zip root (used for ``_lineage/vehicle.parquet``, etc.). Paths are
      relative to ``root`` and can contain subdirectories.
    """
    ts = timestamp or datetime.now()
    ts_str = ts.strftime("%Y%m%d_%H%M")
    client_slug = _safe_client_slug(client_name)
    root = f"revio_import_{client_slug}_{ts_str}"

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # 1. Master workbook — always (now multi-schema).
        master_bytes = build_master_xlsx(
            client_name=client_name,
            vehicle_df=vehicle_df,
            vehicle_fleet_mapping=vehicle_fleet_mapping,
            contract_df=contract_df,
            contract_orphan_df=contract_orphan_df,
            contract_fleet_mapping=contract_fleet_mapping,
        )
        zf.writestr(f"{root}/onboarding_complet.xlsx", master_bytes)

        # 2. Vehicles CSVs.
        if vehicle_df is not None and not vehicle_df.empty:
            zf.writestr(
                f"{root}/vehicles/vehicles_tous.csv",
                revio_csv_bytes(vehicle_df, "vehicle"),
            )
            if vehicle_fleet_mapping is not None and vehicle_fleet_mapping.is_active:
                buckets = split_df_by_fleet(vehicle_df, vehicle_fleet_mapping)
                for fleet_name in _fleet_iter_order(buckets.keys()):
                    sub = buckets[fleet_name]
                    if sub.empty:
                        continue
                    slug = slugify_fleet_name(fleet_name) or "fleet"
                    zf.writestr(
                        f"{root}/vehicles/vehicles_{slug}.csv",
                        revio_csv_bytes(sub, "vehicle"),
                    )

        # 3. Contracts CSVs.
        if contract_df is not None and not contract_df.empty:
            zf.writestr(
                f"{root}/contracts/contracts_tous.csv",
                revio_csv_bytes(contract_df, "contract"),
            )
            if contract_fleet_mapping is not None and contract_fleet_mapping.is_active:
                c_buckets = _split_contracts_by_fleet(contract_df, contract_fleet_mapping)
                for fleet_name in _fleet_iter_order(c_buckets.keys()):
                    sub = c_buckets[fleet_name]
                    if sub.empty:
                        continue
                    slug = slugify_fleet_name(fleet_name) or "fleet"
                    zf.writestr(
                        f"{root}/contracts/contracts_{slug}.csv",
                        revio_csv_bytes(sub, "contract"),
                    )

        # Contract orphans (raw columns, not Revio header).
        if contract_orphan_df is not None and not contract_orphan_df.empty:
            buf_orph = io.StringIO()
            contract_orphan_df.to_csv(buf_orph, index=False)
            zf.writestr(
                f"{root}/contracts/contracts_orphelins.csv",
                buf_orph.getvalue().encode("utf-8-sig"),
            )

        # 4. Report / errors xlsx — passthroughs.
        if report_xlsx_bytes:
            zf.writestr(f"{root}/rapport.xlsx", report_xlsx_bytes)
        if contract_errors_xlsx_bytes:
            zf.writestr(f"{root}/contracts_errors.xlsx", contract_errors_xlsx_bytes)

        # 5. Extra files (e.g. lineage sidecars _lineage/vehicle.parquet).
        if extra_files:
            for rel_path, content in extra_files.items():
                if content is None:
                    continue
                safe_rel = str(rel_path).lstrip("/\\")
                zf.writestr(f"{root}/{safe_rel}", content)

    filename = f"{root}.zip"
    return buf.getvalue(), filename
