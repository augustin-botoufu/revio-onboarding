"""Revio Onboarding Tool - Streamlit app.

Run locally:
    pip install -r requirements.txt
    streamlit run app.py
"""

from __future__ import annotations

import os
from pathlib import Path

import pandas as pd
import streamlit as st

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from src.detectors import detect, label_for, SOURCE_TYPE_LABELS
from src.llm_mapper import propose_mapping
from src.pipeline import (
    SourceFile,
    load_tabular,
    merge_per_schema,
    validate,
    apply_mapping,
    merge_engine_sources,
)
from src.output_writer import build_zip, split_by_fleet
from src.schemas import SCHEMAS, header_for
from src import rules_engine
from src import rules_io
from src import contract_engine as ce
from src import pdf_parser as pdfp
from src import unknown_columns as unkcol
from src import learned_patterns as lp
from src import github_sync as gh
from src import value_mappings as vm
from src import ai_normalization as ain
from src import fleet_segmentation as fseg
from src import zip_writer as zw
from src.excel_report import build_report_xlsx
from src import auth as rv_auth
from src.session_reset import reset_import_state


# Mapping between the `detect()` source_type and the YAML `source` slug.
# Sources not listed are unavailable to the rules engine (templates, etc.).
DETECTOR_TO_YAML_SLUG = {
    "api_plaques": "api_plaques",
    "ayvens_etat_parc": "ayvens_etat_parc",
    "ayvens_aen": "ayvens_aen",
    "ayvens_tvs": "ayvens_tvs",
    "ayvens_and": "ayvens_and",
    "arval_uat": "arval_uat",
    "arval_aen": "arval_aen",
    "arval_tvu": "arval_tvu",
    "arval_and": "arval_and",
    "client_vehicle": "client_file",
}

# All YAML slugs the engine understands (for the override dropdown).
ENGINE_SOURCE_SLUGS = [
    "api_plaques",
    "ayvens_etat_parc",
    "ayvens_aen",
    "ayvens_tvs",
    "ayvens_and",
    "ayvens_pneus",
    "arval_uat",
    "arval_aen",
    "arval_tvu",
    "arval_and",
    "arval_pneus",
    "autre_loueur_etat_parc",
    "autre_loueur_aen",
    "autre_loueur_tvs",
    "autre_loueur_and",
    "autre_loueur_pneus",
    "assurance_externe",
    "client_file",
    # --- Contract-only sources (factures PDF) ---
    "arval_facture_pdf",
    "ayvens_facture_pdf",
    "autre_loueur_facture_pdf",
]

# Human-readable label + icon for each source slug, for the grouped cards
# in the Moteur page. The tuple is (emoji, label, group_order) — group_order
# drives the rendering order of the cards so the page reads top-to-bottom in
# a predictable way (plaques first, then by lessor, then catch-alls at the end).
SLUG_DISPLAY: dict[str, tuple[str, str, int]] = {
    "api_plaques":             ("🔢", "API Plaques",                  10),
    "ayvens_etat_parc":        ("🚗", "Ayvens — État de parc",        20),
    "ayvens_aen":              ("🚗", "Ayvens — Avis d'échéance",     21),
    "ayvens_tvs":              ("🚗", "Ayvens — TVS",                  22),
    "ayvens_and":              ("🚗", "Ayvens — Avis non dépôt",       23),
    "ayvens_pneus":            ("🚗", "Ayvens — Pneus",                24),
    "arval_uat":               ("🚙", "Arval — UAT",                   30),
    "arval_aen":               ("🚙", "Arval — Avis d'échéance",       31),
    "arval_tvu":               ("🚙", "Arval — TVU",                   32),
    "arval_and":               ("🚙", "Arval — Avis non dépôt",        33),
    "arval_pneus":             ("🚙", "Arval — Pneus",                 34),
    "autre_loueur_etat_parc":  ("📘", "Autre loueur — État de parc",   40),
    "autre_loueur_aen":        ("📘", "Autre loueur — Avis d'échéance",41),
    "autre_loueur_tvs":        ("📘", "Autre loueur — TVS",            42),
    "autre_loueur_and":        ("📘", "Autre loueur — Avis non dépôt", 43),
    "autre_loueur_pneus":      ("📘", "Autre loueur — Pneus",          44),
    "assurance_externe":       ("🛡️", "Assurance externe",             50),
    "arval_facture_pdf":       ("📄", "Arval — Facture (PDF)",          60),
    "ayvens_facture_pdf":      ("📄", "Ayvens — Facture (PDF)",         61),
    "autre_loueur_facture_pdf":("📄", "Autre loueur — Facture (PDF)",   62),
    "client_file":             ("👤", "Fichier client",                90),
}


def slug_display(slug: str) -> tuple[str, str, int]:
    """Return (emoji, label, order) for a slug. Unknown slugs fall back gracefully."""
    return SLUG_DISPLAY.get(slug, ("📄", slug, 99))


# Fields that need a manual column mapping at upload time. Used for slugs
# whose column names CAN'T be pre-declared in vehicle.yml, because the
# format isn't known in advance:
#   - client_file = client's free-form vehicle spreadsheet
#   - autre_loueur_etat_parc = unknown lessor export
# For both slugs, the user (or the AI) fills the source column name per
# Revio field at runtime. The rules engine picks up these overrides as a
# fallback when the YAML rule has no `column:` key.
MANUAL_MAPPABLE_FIELDS = [
    "registrationPlate",
    "usage",
    "parcEntryAt",
    "registrationIssueCountryCode",
    "brand",
    "model",
    "variant",
    "motorisation",
    "co2gKm",
    "registrationIssueDate",
    "registrationVin",
    "registrationFiscalPower",
]

# Alias kept so other parts of the codebase that reference the old name
# keep working. Safe to remove once all call sites are migrated.
CLIENT_FILE_MAPPABLE_FIELDS = MANUAL_MAPPABLE_FIELDS

# Slugs that require a manual column mapping (LLM-assisted or hand-picked)
# before the rules engine can read anything from the file. All OTHER slugs
# (ayvens_*, arval_*, api_plaques, …) have their column names baked into
# vehicle.yml and don't need this step.
SLUGS_NEEDING_MANUAL_MAPPING = {"client_file", "autre_loueur_etat_parc"}


# Contract-side mappable fields — expected in the client_file for contracts.
# Used by the "colonne non identifiée" UI when the engine can't resolve a
# mandatory field on its own.
MANUAL_MAPPABLE_FIELDS_CONTRACT = [
    "plate", "number", "startDate", "endDate", "durationMonths",
    "contractedMileage", "maxMileage", "extraKmPrice",
    "vehicleValue", "batteryValue", "totalPrice",
    "civilLiabilityPrice", "allRisksPrice", "maintenancePrice",
    "partnerId", "isHT",
]


# ========== Page config ==========
st.set_page_config(
    page_title="Revio — Onboarding",
    page_icon=str(Path(__file__).parent / "src" / "assets" / "logo.svg")
    if (Path(__file__).parent / "src" / "assets" / "logo.svg").exists()
    else "🚗",
    layout="wide",
    initial_sidebar_state="expanded",
)


# ========== Password gate ==========
# Blocks the rest of the app until the user enters the shared team password
# (configured in Streamlit Cloud → Settings → Secrets as `app_password`).
# If no password is set, runs in "dev mode" with auth disabled.
rv_auth.require_password()


# ========== Branding (logo in the top-left, replaces the old sidebar title) ==========
_ASSET_DIR = Path(__file__).parent / "src" / "assets"
_LOGO_WORDMARK = _ASSET_DIR / "logo_wordmark.svg"  # full "R + revio"
_LOGO_ICON = _ASSET_DIR / "logo.svg"               # just the R (sidebar collapsed)
if _LOGO_WORDMARK.exists():
    try:
        st.logo(
            str(_LOGO_WORDMARK),
            icon_image=str(_LOGO_ICON) if _LOGO_ICON.exists() else None,
            size="large",
        )
    except TypeError:
        # Streamlit <1.37: `size` kwarg not supported.
        st.logo(
            str(_LOGO_WORDMARK),
            icon_image=str(_LOGO_ICON) if _LOGO_ICON.exists() else None,
        )


# ========== Custom style (typography + component polish) ==========
def _inject_custom_style() -> None:
    """Give the app a modern, editorial look on top of the base theme.

    - Inter (Google Fonts) as the primary typeface.
    - Softer borders, unified 8-10px radius, tightened spacing.
    - Hides Streamlit's default top chrome (menu, deploy button, footer).
    Additive: the .streamlit/config.toml theme stays the source of truth
    for colors — this stylesheet only refines typography + components.
    """
    st.markdown(
        """
        <style>
          @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

          html, body, .stApp, [data-testid="stAppViewContainer"], [data-testid="stSidebar"] {
              font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Helvetica, Arial, sans-serif !important;
          }

          /* Hide Streamlit's default top chrome */
          #MainMenu { visibility: hidden; }
          footer { visibility: hidden; }
          header[data-testid="stHeader"] { background: transparent; height: 0; }

          .main .block-container {
              padding-top: 2.25rem;
              padding-bottom: 4rem;
              max-width: 1240px;
          }

          /* Typography */
          h1, h2, h3, h4, h5, h6 {
              font-family: 'Inter', sans-serif !important;
              font-weight: 600 !important;
              letter-spacing: -0.015em;
              color: #0F172A;
          }
          h1 { font-size: 1.875rem !important; line-height: 1.2; }
          h2 { font-size: 1.375rem !important; margin-top: 1.75rem !important; }
          h3 { font-size: 1.0625rem !important; margin-top: 1.25rem !important; color: #1E293B !important; }
          [data-testid="stCaptionContainer"], .stCaption { color: #64748B !important; font-size: 0.85rem; }

          /* Buttons */
          .stButton > button,
          [data-testid="stFormSubmitButton"] > button,
          [data-testid="stDownloadButton"] > button {
              border-radius: 8px !important;
              border: 1px solid #E2E8F0 !important;
              font-weight: 500 !important;
              padding: 0.5rem 1rem !important;
              transition: all 0.15s ease !important;
              box-shadow: 0 1px 2px rgba(15, 23, 42, 0.04);
          }
          .stButton > button:hover,
          [data-testid="stFormSubmitButton"] > button:hover,
          [data-testid="stDownloadButton"] > button:hover {
              border-color: #CBD5E1 !important;
              background: #F8FAFC !important;
          }
          .stButton > button[kind="primary"],
          [data-testid="stFormSubmitButton"] > button[kind="primary"],
          [data-testid="stDownloadButton"] > button[kind="primary"] {
              background: #0F172A !important;
              color: #FFFFFF !important;
              border-color: #0F172A !important;
          }
          .stButton > button[kind="primary"]:hover {
              background: #1E293B !important;
              border-color: #1E293B !important;
          }

          /* Inputs */
          .stTextInput input, .stTextArea textarea,
          .stSelectbox div[data-baseweb="select"] > div,
          .stNumberInput input, .stDateInput input {
              border-radius: 8px !important;
              border-color: #E2E8F0 !important;
          }

          /* Expanders */
          [data-testid="stExpander"] {
              border: 1px solid #E2E8F0 !important;
              border-radius: 10px !important;
              box-shadow: 0 1px 2px rgba(15, 23, 42, 0.03);
              background: #FFFFFF;
          }
          [data-testid="stExpander"] summary { font-weight: 500 !important; }

          /* Alerts */
          div[data-testid="stAlert"] {
              border-radius: 10px !important;
              padding: 0.875rem 1rem !important;
              font-size: 0.9rem;
              border: 1px solid rgba(15, 23, 42, 0.06);
          }

          /* Tabs */
          [data-testid="stTabs"] [role="tablist"] {
              border-bottom: 1px solid #E2E8F0;
              gap: 0.125rem;
          }
          [data-testid="stTabs"] [role="tab"] {
              font-weight: 500;
              color: #64748B;
              padding: 0.625rem 1rem;
          }
          [data-testid="stTabs"] [role="tab"][aria-selected="true"] { color: #0F172A; }

          /* Dataframes */
          [data-testid="stDataFrame"], [data-testid="stTable"] {
              border-radius: 10px;
              overflow: hidden;
              border: 1px solid #E2E8F0;
          }

          /* Metrics */
          [data-testid="stMetric"] {
              background: #FFFFFF;
              border: 1px solid #E2E8F0;
              border-radius: 10px;
              padding: 1rem 1.25rem;
              box-shadow: 0 1px 2px rgba(15, 23, 42, 0.03);
          }

          /* File uploader */
          [data-testid="stFileUploader"] section {
              border-radius: 10px;
              border: 1px dashed #CBD5E1;
              background: #F8FAFC;
          }

          /* Sidebar polish */
          [data-testid="stSidebar"] {
              background: #FAFBFC;
              border-right: 1px solid #E2E8F0;
          }
          [data-testid="stSidebar"] h3 {
              font-size: 0.72rem !important;
              text-transform: uppercase;
              letter-spacing: 0.08em;
              color: #94A3B8 !important;
              font-weight: 600 !important;
              margin-top: 1.5rem !important;
              margin-bottom: 0.5rem !important;
          }
          [data-testid="stSidebar"] hr { margin: 1rem 0 !important; border-color: #E2E8F0; }
          [data-testid="stSidebar"] [data-baseweb="radio"] { padding: 0.25rem 0; }
          [data-testid="stSidebar"] [data-baseweb="radio"] label {
              font-size: 0.93rem; font-weight: 500;
          }

          /* st.logo sizing */
          [data-testid="stLogo"], [data-testid="stSidebarHeader"] img {
              height: 36px !important; width: auto !important;
          }

          /* Divider */
          hr { border-color: #E2E8F0; margin: 1.25rem 0; }

          /* Popover trigger */
          [data-testid="stPopover"] button {
              border-radius: 8px;
              background: #FFFFFF;
              border: 1px solid #E2E8F0;
          }

          /* Sidebar nav: full-width, left-aligned buttons (replaces radio) */
          [data-testid="stSidebar"] .stButton > button {
              justify-content: flex-start !important;
              text-align: left !important;
              padding-left: 0.875rem !important;
              padding-right: 0.875rem !important;
              font-weight: 500 !important;
              box-shadow: none !important;
              background: transparent !important;
              border: 1px solid transparent !important;
              color: #475569 !important;
              margin-bottom: 0.125rem !important;
          }
          [data-testid="stSidebar"] .stButton > button:hover {
              background: #F1F5F9 !important;
              color: #0F172A !important;
              border-color: transparent !important;
          }
          [data-testid="stSidebar"] .stButton > button[kind="primary"] {
              background: #0F172A !important;
              color: #FFFFFF !important;
              border-color: #0F172A !important;
          }
          [data-testid="stSidebar"] .stButton > button[kind="primary"]:hover {
              background: #1E293B !important;
              color: #FFFFFF !important;
          }

          /* Home page: action cards */
          .home-hero {
              background: linear-gradient(135deg, #F8FAFC 0%, #EEF2F7 100%);
              border: 1px solid #E2E8F0;
              border-radius: 16px;
              padding: 2.25rem 2.5rem;
              margin-bottom: 2rem;
          }
          .home-hero h1 { margin-top: 0 !important; margin-bottom: 0.35rem !important; }
          .home-hero p { color: #475569; font-size: 1rem; max-width: 640px; margin: 0; }

          /* Stepper: horizontal progress in main page (replaces sidebar Étapes) */
          .rv-stepper {
              display: flex;
              gap: 0.5rem;
              margin: 0 0 1.5rem 0;
              padding: 0.5rem 0;
              overflow-x: auto;
          }
          .rv-step {
              display: flex;
              align-items: center;
              gap: 0.5rem;
              padding: 0.5rem 0.875rem;
              border-radius: 999px;
              border: 1px solid #E2E8F0;
              background: #FFFFFF;
              font-size: 0.85rem;
              font-weight: 500;
              color: #64748B;
              white-space: nowrap;
              cursor: default;
          }
          .rv-step.is-active {
              background: #0F172A;
              border-color: #0F172A;
              color: #FFFFFF;
          }
          .rv-step.is-done {
              background: #F1F5F9;
              border-color: #CBD5E1;
              color: #334155;
          }
          .rv-step .rv-step-num {
              display: inline-flex;
              align-items: center;
              justify-content: center;
              width: 22px; height: 22px;
              border-radius: 50%;
              background: #E2E8F0;
              color: #334155;
              font-size: 0.75rem;
              font-weight: 600;
          }
          .rv-step.is-active .rv-step-num { background: rgba(255,255,255,0.2); color: #FFFFFF; }
          .rv-step.is-done .rv-step-num { background: #CBD5E1; color: #0F172A; }

          /* Small "pill" badge (session overrides indicator, etc.) */
          .rv-pill {
              display: inline-flex;
              align-items: center;
              gap: 0.35rem;
              padding: 0.25rem 0.625rem;
              border-radius: 999px;
              background: #FEF3C7;
              color: #92400E;
              font-size: 0.78rem;
              font-weight: 500;
              border: 1px solid #FDE68A;
          }
        </style>
        """,
        unsafe_allow_html=True,
    )


_inject_custom_style()


# ========== Session state init ==========
def _init_state():
    defaults = {
        "sources": {},              # {key: SourceFile}
        "client_name": "",
        "user_instructions": "",
        "llm_proposals": {},        # {key: {target_field: source_col}}
        "fleet_mapping": {},        # {agency_code: fleet_name}
        "step": 1,
        "mode": "home",             # "home" | "classic" | "engine" | "rules"
        # --- engine mode ---
        "engine_files": {},         # {filename: {"df": DataFrame, "slug": str, "detected": str}}
        "engine_overrides": {},     # {(file_key, field): source_col} — per-FILE mapping
        "engine_result": None,      # EngineResult (dataclass) or None
        # --- value normalization (Jalon 2.7) ---
        # Cached once per session — reused across engine runs, AI fallback, and UI.
        "value_mappings": None,     # dict[str, dict[str, vm.ValueMapping]] or None
        "ai_fallback_report": None, # AIFallbackReport of the last engine run
        # --- fleet segmentation (Jalon 3.0) ---
        # Session-scoped FleetMapping (or None). Lives here so the popup
        # can be re-opened with the previous choices preselected.
        "engine_fleet_mapping": None,  # fleet_segmentation.FleetMapping or None
        # Scratch area used by the @st.dialog while the user edits the
        # mapping. Flushed to engine_fleet_mapping only on « Valider ».
        "fleet_dialog_draft": None,    # dict with keys: file_key, column, raw_to_fleet
        # --- rules editor (session-scoped priority overrides) ---
        # Shape: {table_slug: {field_name: [source_slug_in_priority_order]}}
        "rules_overrides": {},
        "rules_active_table": "vehicle",
        # --- Contract engine (Jalon 4.2.2) ---
        # Contract shares `engine_files` with the Vehicle tab — a single uploader
        # populates both tables, and the Contract engine simply ignores slugs
        # that aren't declared in contract.yml. Only result + user picks are
        # per-table.
        "contract_result": None,            # ContractEngineResult or None
        # Resolved answers from the "colonne non identifiée" UI. Kept in memory
        # until the user validates — then persisted via register_learned_column.
        "contract_unknown_resolved": {},    # {(source_slug, field_name): column}
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


_init_state()


# ========== Sidebar ==========
# Structure:
#   - st.logo() — handled above, renders the R + "revio" wordmark
#   - Navigation: 4 button-style menu items (Accueil / Import classique / Moteur / Règles)
#   - Contexte: client name + API key status (only on onboarding modes)
#   - Instructions spéciales: collapsed expander (only on onboarding modes)
NAV_ITEMS = [
    ("home",    "🏠",  "Accueil"),
    ("classic", "📥",  "Import — Flow classique"),
    ("engine",  "🧪",  "Import — Moteur de règles"),
    ("rules",   "⚙️",  "Règles d'import"),
    ("normal",  "🧭",  "Normalisation"),
]

with st.sidebar:
    st.caption("Outil interne d'onboarding — génération des fichiers d'import.")

    st.markdown("### Navigation")
    for slug, icon, label in NAV_ITEMS:
        active = st.session_state.mode == slug
        if st.button(
            f"{icon}  {label}",
            key=f"nav_{slug}",
            use_container_width=True,
            type="primary" if active else "secondary",
        ):
            if not active:
                st.session_state.mode = slug
                st.rerun()

    # Session-scoped rule overrides indicator — small pill badge, not a caption.
    _nb_overrides = rules_io.count_active_overrides(st.session_state.get("rules_overrides"))
    if _nb_overrides > 0:
        st.markdown(
            f"<div style='margin-top:0.5rem'><span class='rv-pill'>✎ {_nb_overrides} règle(s) personnalisée(s)</span></div>",
            unsafe_allow_html=True,
        )

    # --- Contexte (only on onboarding modes) ---
    if st.session_state.mode in ("classic", "engine"):
        st.markdown("---")
        st.markdown("### Contexte")
        st.session_state.client_name = st.text_input(
            "Nom du client",
            value=st.session_state.client_name,
            placeholder="ex. YSEIS",
            help="Utilisé pour nommer le dossier de sortie.",
        )

        # API key: (1) Streamlit Cloud secrets, (2) local .env, (3) manual paste.
        api_key_env = os.environ.get("ANTHROPIC_API_KEY", "")
        if not api_key_env:
            try:
                api_key_env = st.secrets.get("ANTHROPIC_API_KEY", "")
                if api_key_env:
                    os.environ["ANTHROPIC_API_KEY"] = api_key_env
            except Exception:
                pass
        if api_key_env:
            st.success("Clé Anthropic détectée ✓")
        else:
            st.warning("Pas de clé API — colle-la pour activer le mapping IA.")
            pasted = st.text_input("Clé Anthropic (sk-ant-...)", type="password")
            if pasted:
                os.environ["ANTHROPIC_API_KEY"] = pasted

        with st.expander("✍️ Instructions spéciales", expanded=False):
            st.caption(
                "Règles en langage naturel qui s'appliquent à tout l'onboarding. "
                "*Ex : Pour ce client, un VP-BR = service. Si la date fin est vide, calcule Date début + Durée.*"
            )
            st.session_state.user_instructions = st.text_area(
                "Instructions",
                value=st.session_state.user_instructions,
                height=140,
                label_visibility="collapsed",
            )

        # GitHub sync diagnostics — lets the user verify their PAT works
        # without having to memorize a real pattern (Jalon 2.5 safety net).
        with st.expander("🔌 Sync GitHub (mémoire)", expanded=False):
            if not gh.is_configured():
                st.caption(
                    "Auto-commit GitHub non activé. Ajoute `GITHUB_TOKEN` et "
                    "`GITHUB_REPO` dans Streamlit Cloud → Settings → Secrets "
                    "pour activer la mémorisation automatique. Sans ça, la "
                    "mémorisation fonctionne toujours mais en copier-coller."
                )
            else:
                if st.button(
                    "🔌 Tester la connexion",
                    key="gh_check_connection",
                    use_container_width=True,
                    help="Lit `learned_patterns.yml` sur GitHub pour vérifier "
                         "que le token et le chemin sont bons. N'écrit rien.",
                ):
                    with st.spinner("Test de la connexion GitHub..."):
                        st.session_state["gh_check_result"] = gh.check_connection()
                result = st.session_state.get("gh_check_result")
                if result:
                    if not result.get("ok"):
                        st.error(f"❌ {result.get('message', 'Échec')}")
                    elif not result.get("file_exists"):
                        st.warning(f"⚠️ {result.get('message')}")
                    else:
                        st.success(f"✅ {result.get('message')}")
                    if result.get("configured"):
                        st.caption(
                            f"Repo : `{result.get('repo')}` · "
                            f"branche : `{result.get('branch')}` · "
                            f"fichier : `{result.get('path')}`"
                        )

    # --- Session actions (always visible, at the bottom of the sidebar) ---
    st.markdown("---")
    # "Nouvel import" — clears uploaded files, mappings, engine result, fleet
    # mapping, … but keeps auth + current nav. No page refresh needed.
    if st.button(
        "🔄 Nouvel import",
        key="btn_reset_import",
        use_container_width=True,
        help="Vide tous les fichiers et mappings en cours pour repartir "
             "d'un onboarding propre, sans rafraîchir la page.",
    ):
        reset_import_state(st.session_state)
        st.session_state["mode"] = "home"
        st.toast("Session remise à zéro — prêt pour un nouvel import.", icon="🔄")
        st.rerun()

    # Déconnexion — only if auth is actually active.
    if not rv_auth.is_auth_disabled():
        if st.button(
            "🔓 Déconnexion",
            key="btn_logout",
            use_container_width=True,
            help="Quitter la session et revenir à l'écran de mot de passe.",
        ):
            rv_auth.logout()
            # Wipe all import state too, so the next user starts fresh.
            reset_import_state(st.session_state)
            st.rerun()
    else:
        st.caption(
            "🔓 Accès libre — définissez `app_password` dans les Secrets "
            "Streamlit Cloud pour restreindre l'accès."
        )


# ========== Home page (landing) ==========
def render_home_page() -> None:
    """Welcoming landing page: hero + 3 action cards pointing to each mode.

    Shown on arrival (mode default = "home") and whenever the user clicks
    the "🏠 Accueil" item in the sidebar.
    """
    # Hero band
    st.markdown(
        """
        <div class="home-hero">
          <h1>Onboarding Revio</h1>
          <p>Transforme les fichiers reçus d'un client (API Plaques, exports loueurs,
          fichier interne) en imports Revio prêts à lancer. Moteur de règles YAML,
          mapping IA assisté, contrôle qualité intégré.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # Primary CTA + quick-access cards
    st.markdown("### Par où veux-tu commencer ?")
    c1, c2, c3 = st.columns(3, gap="medium")

    with c1:
        with st.container(border=True):
            st.markdown("#### 🧪 Moteur de règles")
            st.caption(
                "Dépose tes fichiers, le moteur applique les règles YAML et "
                "produit l'import Vehicle. **Recommandé.**"
            )
            if st.button("Ouvrir le moteur", key="home_cta_engine", use_container_width=True, type="primary"):
                st.session_state.mode = "engine"
                st.rerun()

    with c2:
        with st.container(border=True):
            st.markdown("#### 📥 Flow classique")
            st.caption(
                "Le flow en 5 étapes : upload → détection → mapping manuel → "
                "découpage flottes → exports."
            )
            if st.button("Ouvrir le flow", key="home_cta_classic", use_container_width=True):
                st.session_state.mode = "classic"
                st.session_state.step = 1
                st.rerun()

    with c3:
        with st.container(border=True):
            st.markdown("#### ⚙️ Règles d'import")
            st.caption(
                "Consulte et ajuste la priorité des sources par champ. "
                "Modifications session-only."
            )
            if st.button("Ouvrir les règles", key="home_cta_rules", use_container_width=True):
                st.session_state.mode = "rules"
                st.rerun()

    # A soft "what's new" hint at the bottom — keeps the page from feeling empty
    # without overselling.
    st.markdown("")
    st.markdown("---")
    st.markdown("##### Astuces")
    a, b, c = st.columns(3, gap="medium")
    with a:
        st.markdown("**Un seul onboarding = un seul client.**")
        st.caption("Dépose tous les fichiers reçus d'un coup, l'outil détecte chaque source automatiquement.")
    with b:
        st.markdown("**Les règles peuvent être surchargées.**")
        st.caption("Sur la page *Règles d'import*, l'ordre de priorité est modifiable pour un onboarding sans toucher la config permanente.")
    with c:
        st.markdown("**Tout est interne.**")
        st.caption("Les fichiers uploadés ne quittent pas l'app Streamlit. Rien n'est persisté côté serveur.")


# ========== Classic-mode stepper (horizontal progress bar) ==========
CLASSIC_STEPS: list[tuple[int, str]] = [
    (1, "Upload"),
    (2, "Détection"),
    (3, "Mapping"),
    (4, "Flottes"),
    (5, "Exports"),
]


def _render_classic_stepper() -> None:
    """Horizontal pill-stepper shown at the top of each classic-mode step.

    Replaces the sidebar radio: users navigate by clicking the pills. Visual
    state: previous steps = "done" (light), current = "active" (dark), future
    = default (white).
    """
    current = int(st.session_state.get("step", 1))

    # Render the visual strip via HTML for a pixel-perfect pill look.
    pills = []
    for n, label in CLASSIC_STEPS:
        cls = "rv-step"
        if n == current:
            cls += " is-active"
        elif n < current:
            cls += " is-done"
        pills.append(
            f'<div class="{cls}"><span class="rv-step-num">{n}</span>{label}</div>'
        )
    st.markdown(f'<div class="rv-stepper">{"".join(pills)}</div>', unsafe_allow_html=True)

    # Jump buttons sit just below so keyboard/click navigation still works
    # (HTML pills aren't clickable in Streamlit; the buttons are the real UX).
    cols = st.columns(len(CLASSIC_STEPS))
    for idx, (n, label) in enumerate(CLASSIC_STEPS):
        with cols[idx]:
            active = n == current
            if st.button(
                f"{n}. {label}",
                key=f"stepper_jump_{n}",
                use_container_width=True,
                type="primary" if active else "secondary",
            ):
                if not active:
                    st.session_state.step = n
                    st.rerun()
    st.markdown("")  # small spacer before the step content


# ========== Engine mode (YAML rules engine, Vehicle only) ==========
def _current_user_email() -> str | None:
    """Best-effort read of the Streamlit-authenticated user's email.

    Used as metadata on learned_patterns entries (created_by) and on GitHub
    commits (committer email). Safe to call outside an auth context — returns
    None if no user info is available.
    """
    try:
        user_obj = getattr(st, "user", None)
        if user_obj is None:
            return None
        if hasattr(user_obj, "get"):
            return user_obj.get("email") or None
        return getattr(user_obj, "email", None) or None
    except Exception:
        return None


def _render_column_preview(df, col: str, max_values: int = 15) -> None:
    """Render a compact preview of a column's values inside a popover.

    Shown when the user clicks the 🔍 button next to a mapping selectbox so
    they can verify what's actually in the source column without going back
    to the file. Keeps it snappy: metrics on top, up to ``max_values`` unique
    samples below, each truncated if longer than 120 chars.
    """
    if col not in df.columns:
        st.warning(f"Colonne `{col}` introuvable dans le fichier.")
        return
    s = df[col]
    total = int(len(s))
    non_null = int(s.notna().sum())
    nulls = total - non_null
    try:
        uniques = (
            s.dropna()
            .astype(str)
            .str.strip()
            .replace("", pd.NA)
            .dropna()
            .unique()
        )
    except Exception:
        uniques = [str(v) for v in s.dropna().unique()]
    n_unique = int(len(uniques))

    st.markdown(f"**Colonne `{col}`**")
    m1, m2, m3 = st.columns(3)
    m1.metric("Valeurs", total)
    m2.metric("Uniques", n_unique)
    m3.metric("Vides", nulls)

    if n_unique == 0:
        st.info("Cette colonne est entièrement vide.")
        return

    shown = list(uniques[:max_values])
    truncated = n_unique > max_values
    st.caption(
        f"Aperçu ({len(shown)}/{n_unique} valeurs uniques)"
        + (" — tronqué" if truncated else "")
        + " :"
    )
    for v in shown:
        v_str = str(v)
        if len(v_str) > 120:
            v_str = v_str[:117] + "..."
        st.markdown(f"- `{v_str}`")


def _render_recognized_card(key: str, info: dict) -> None:
    """Render one compact card in the 'Fichiers reconnus automatiquement' zone.

    Read-only view by default (title + summary + 👁️ Voir popover). Two
    action buttons on the right: « ✏️ Modifier » flips the file into edit
    mode (it reappears in the full mapping UI below on next rerun) and
    « 🗑️ » removes the pattern from GitHub (same call as the in-card
    delete button, short-circuited here for convenience).
    """
    learned_id = info.get("learned_match_id")
    learned_hint = info.get("learned_match_hint")
    mapping_applied = int(info.get("learned_mapping_applied_count") or 0)
    slug_label = slug_display(info["slug"])[1]
    label = info["filename"] + (
        f" [{info['sheet_name']}]" if info.get("sheet_name") else ""
    )
    current_mapping = {
        fld: src_col
        for (fk, fld), src_col in st.session_state.engine_overrides.items()
        if fk == key and src_col
    }

    with st.container(border=True):
        c_title, c_view, c_edit, c_del = st.columns([6, 2, 2, 1])
        with c_title:
            st.markdown(f"**🧠 {label}** → {slug_label}")
            hint_suffix = f" · loueur *{learned_hint}*" if learned_hint else ""
            st.caption(
                f"Pattern `{learned_id}`{hint_suffix} · "
                f"{mapping_applied} champ(s) mappé(s)"
            )
        with c_view:
            with st.popover(
                "👁️ Voir",
                use_container_width=True,
                help="Voir le mapping actif sans le modifier",
            ):
                st.markdown(f"**Mapping actif — `{label}`**")
                if not current_mapping:
                    st.caption("Aucun champ mappé.")
                else:
                    for fld, src in current_mapping.items():
                        st.markdown(f"- **{fld}** ← `{src}`")
        with c_edit:
            if st.button(
                "✏️ Modifier",
                key=f"engine_edit_recognized_btn_{key}",
                use_container_width=True,
                help="Rouvre la carte de mapping complète pour ajuster les champs",
            ):
                st.session_state[f"engine_edit_recognized_{key}"] = True
                st.rerun()
        with c_del:
            del_disabled = not (gh.is_configured() and learned_id)
            if st.button(
                "🗑️",
                key=f"engine_del_recognized_btn_{key}",
                use_container_width=True,
                disabled=del_disabled,
                help=(
                    "Configure GitHub dans la sidebar pour activer la suppression."
                    if del_disabled
                    else f"Supprimer le pattern `{learned_id}` sur GitHub"
                ),
            ):
                try:
                    resp = gh.delete_pattern(
                        learned_id,
                        author_email=_current_user_email(),
                    )
                    if resp.get("skipped"):
                        st.warning(
                            f"Le pattern `{learned_id}` n'existait "
                            "plus sur GitHub — rien à supprimer."
                        )
                    else:
                        st.success(
                            f"✅ Pattern `{learned_id}` supprimé sur "
                            "GitHub. Streamlit va redéployer dans ~1 min — "
                            "recharge la page après."
                        )
                except gh.GitHubSyncError as e:
                    st.error(f"❌ {e.user_message}")


def _render_normalizations_report(result, ai_report) -> None:
    """Report tab for Jalon 2.7 value-level normalization.

    Shows 3 sub-sections:
    - ✅ Valeurs normalisées via le dictionnaire (cache hits — validated)
    - 🤖 Proposées par l'IA (pending — click to validate)
    - ❓ Valeurs inconnues (cache + IA miss)

    Nothing here if the engine didn't touch any enum field (no noise when
    the user's files are already clean).
    """
    if result is None:
        return
    hits = dict(result.value_mapping_hits or {})
    unresolved = dict(result.unresolved_enums or {})
    # After AI fallback, resolved cells have been removed from unresolved
    # (they moved to hits as pending) — recompute here to be safe in case
    # the AI report is stale.
    if not hits and not unresolved and (ai_report is None or not ai_report.errors):
        return

    st.markdown("#### 🧭 Normalisations")
    st.caption(
        "Les valeurs brutes (VP/VU, Essence, Hybride-diesel, Oui/Non…) qui ont été "
        "transposées vers les valeurs canoniques Revio (private/utility/service, "
        "diesel/gas/hybrid/electric, TRUE/FALSE…). "
        "Les propositions en attente sont à valider dans la page « 🧭 Normalisation »."
    )

    # Split hits by status so we can highlight the pending ones that need review.
    validated_hits: list[dict] = []
    pending_hits: list[dict] = []
    for (plate, fname), (raw, target, status) in hits.items():
        row = {
            "plaque": plate,
            "champ": fname,
            "valeur source": raw,
            "→ valeur Revio": target,
        }
        if status == vm.STATUS_PENDING:
            pending_hits.append(row)
        else:
            validated_hits.append(row)

    col_a, col_b, col_c = st.columns(3)
    col_a.metric("✅ Dictionnaire", len(validated_hits))
    col_b.metric("🤖 IA (en attente)", len(pending_hits))
    col_c.metric("❓ Inconnues", len(unresolved))

    if ai_report and ai_report.errors:
        for (schema, fname), msg in ai_report.errors.items():
            st.warning(f"IA indisponible pour `{schema}.{fname}` — {msg}")

    if pending_hits:
        st.markdown("**🤖 Propositions IA — à valider**")
        st.caption(
            "Appliquées automatiquement pour ne pas vider la cellule. "
            "Valide / corrige depuis la page « 🧭 Normalisation » (sidebar)."
        )
        st.dataframe(
            pd.DataFrame(pending_hits),
            use_container_width=True, hide_index=True,
        )

    if unresolved:
        st.markdown("**❓ Valeurs inconnues — ni règle ni IA**")
        unresolved_rows = [
            {"plaque": plate, "champ": fname,
             "schéma": schema, "valeur": raw}
            for (plate, fname), (schema, raw) in unresolved.items()
        ]
        st.dataframe(
            pd.DataFrame(unresolved_rows),
            use_container_width=True, hide_index=True,
        )
        st.info(
            "Ces cellules conservent leur valeur brute en sortie. "
            "Ajoute une correspondance manuelle depuis la page « 🧭 Normalisation »."
        )

    if validated_hits and not pending_hits and not unresolved:
        with st.expander(f"✅ {len(validated_hits)} valeur(s) normalisée(s) sans intervention"):
            st.dataframe(
                pd.DataFrame(validated_hits),
                use_container_width=True, hide_index=True,
            )


def _render_manual_mapping_section(engine_files: dict) -> None:
    """Render the column-mapping UI for slugs whose columns aren't baked in YAML.

    Used for `client_file` and any `autre_loueur_etat_parc` file (unknown
    lessor export). For each such file we offer a Claude-powered
    auto-mapping button plus manual selectboxes per Revio field. Overrides
    are stored in ``st.session_state.engine_overrides`` with the key
    ``(file_key, field_name)`` — per-FILE, not per-slug, so two files
    sharing a slug (e.g. two different `autre_loueur_etat_parc` exports
    with totally different column headers) can coexist without overwriting
    each other's mapping. At run time, ``merge_engine_sources`` applies
    each file's mapping before concat by copying source columns to
    canonical ``__map__<field>`` names, then the rules engine reads those
    canonical columns via ``(slug, field) → "__map__<field>"`` in
    ``manual_column_overrides``.

    When the user clicks the IA button we update both the overrides dict
    AND the selectbox widget state keys directly, then call ``st.rerun()``
    so the selectboxes reflect the new values on the next render (widget
    state takes precedence over the ``index=`` argument once it exists).

    If no file in ``engine_files`` needs manual mapping, we show the
    existing fallback info about the parc de référence.
    """
    files_to_map = [
        (key, info)
        for key, info in engine_files.items()
        if info.get("slug") in SLUGS_NEEDING_MANUAL_MAPPING
    ]

    if not files_to_map:
        st.info(
            "Aucun fichier à mapper manuellement (pas de `client_file` ni "
            "d'`autre_loueur_etat_parc`) → le moteur prendra l'union des plaques "
            "loueurs comme parc de référence (un avertissement apparaîtra dans "
            "les issues)."
        )
        return

    # Partition: fichiers reconnus par la mémoire ET non ré-ouverts pour
    # édition ET sans pattern en attente de confirmation → zone compacte
    # en haut. Les autres → pleine UI de mapping.
    recognized_auto: list = []
    needs_action: list = []
    for _key, _info in files_to_map:
        _lid = _info.get("learned_match_id")
        _applied = int(_info.get("learned_mapping_applied_count") or 0)
        _edit_mode = st.session_state.get(f"engine_edit_recognized_{_key}", False)
        _has_pending = (
            st.session_state.get(f"engine_pending_pattern_{_key}") is not None
        )
        if _lid and _applied > 0 and not _edit_mode and not _has_pending:
            recognized_auto.append((_key, _info))
        else:
            needs_action.append((_key, _info))

    st.markdown("#### Mapping des colonnes")
    if recognized_auto and not needs_action:
        st.caption(
            "Tous tes fichiers ont été reconnus automatiquement grâce à la "
            "mémoire 🧠. Tu peux lancer le moteur directement, ou rouvrir un "
            "fichier ci-dessous si tu veux ajuster son mapping."
        )
    elif recognized_auto:
        st.caption(
            "Certains fichiers ont été reconnus automatiquement (zone repliée "
            "juste en dessous). Les autres attendent un mapping manuel."
        )
    else:
        st.caption(
            "Pour les fichiers dont le format n'est pas connu à l'avance "
            "(`client_file`, `autre_loueur_etat_parc`), on indique quelle "
            "colonne source correspond à chaque champ Revio. Utilise le "
            "bouton IA pour une proposition automatique, puis révise au "
            "besoin. Les champs non mappés seront ignorés pour ce fichier."
        )

    # ---- Zone compacte : fichiers reconnus automatiquement ----
    if recognized_auto:
        n_reco = len(recognized_auto)
        with st.expander(
            f"✅ Fichiers reconnus automatiquement ({n_reco}) — aucune action requise",
            expanded=False,
        ):
            st.caption(
                "Ces fichiers ont été reconnus par la mémoire 🧠. "
                "« 👁️ Voir » affiche le mapping appliqué, "
                "« ✏️ Modifier » rouvre la carte complète, "
                "« 🗑️ » supprime le pattern sur GitHub."
            )
            for _rkey, _rinfo in recognized_auto:
                _render_recognized_card(_rkey, _rinfo)

    # ---- Pleine UI : fichiers à mapper (ou rouverts pour édition) ----
    if not needs_action:
        return

    if recognized_auto:
        st.markdown("#### ✏️ Fichiers à mapper")

    for key, info in needs_action:
        df = info["df"]
        slug = info["slug"]
        slug_label = slug_display(slug)[1]
        cols = [""] + [str(c) for c in df.columns]
        label = info["filename"] + (
            f" [{info['sheet_name']}]" if info.get("sheet_name") else ""
        )
        # Memory badge: if this file matched a learned pattern at upload
        # and its mapping was auto-applied, surface it in the expander
        # title so the user sees why the selectboxes are pre-filled.
        learned_id = info.get("learned_match_id")
        learned_hint = info.get("learned_match_hint")
        mapping_applied = int(info.get("learned_mapping_applied_count") or 0)
        memory_badge = (
            f" · 🧠 mémoire ({mapping_applied} champ{'s' if mapping_applied > 1 else ''})"
            if learned_id and mapping_applied > 0
            else ""
        )

        edit_mode = st.session_state.get(f"engine_edit_recognized_{key}", False)

        with st.expander(
            f"🔗 {label} → {slug_label} — mapping des champs{memory_badge}",
            expanded=True,
        ):
            if edit_mode:
                ce_left, ce_right = st.columns([5, 1])
                with ce_left:
                    st.caption(
                        "Mode édition : tu modifies le mapping actif. "
                        "Clique sur « Refermer » pour revenir à la zone compacte sans rien toucher."
                    )
                with ce_right:
                    if st.button(
                        "↩️ Refermer",
                        key=f"engine_exit_edit_{key}",
                        use_container_width=True,
                    ):
                        st.session_state.pop(f"engine_edit_recognized_{key}", None)
                        st.rerun()

            if learned_id and mapping_applied > 0:
                st.caption(
                    f"🧠 Format reconnu : `{learned_id}`"
                    + (f" · loueur *{learned_hint}*" if learned_hint else "")
                    + f" — {mapping_applied} champ(s) pré-rempli(s) depuis la mémoire. "
                    "Tu peux quand même relancer l'IA ou ajuster à la main."
                )

            c1, c2 = st.columns([1, 2])
            with c1:
                if st.button(
                    "🤖 Proposer un mapping (IA)",
                    key=f"engine_llm_{key}",
                    use_container_width=True,
                ):
                    with st.spinner("Claude analyse les colonnes..."):
                        result = propose_mapping(
                            df,
                            "vehicle",
                            st.session_state.get("user_instructions", ""),
                        )
                    if "_error" in result:
                        st.error(result["_error"])
                    else:
                        proposed = result.get("mapping", {})
                        notes = result.get("_notes", [])
                        debug = result.get("_debug", {})
                        valid_cols = {str(c) for c in df.columns}
                        nb_mapped = 0
                        for field_name in MANUAL_MAPPABLE_FIELDS:
                            src_col = proposed.get(field_name)
                            widget_key = (
                                f"engine_override_{slug}_{field_name}_{key}"
                            )
                            override_key = (key, field_name)
                            if src_col and src_col in valid_cols:
                                st.session_state.engine_overrides[
                                    override_key
                                ] = src_col
                                st.session_state[widget_key] = src_col
                                nb_mapped += 1
                            else:
                                # Clear any previous value so the selectbox
                                # falls back to "" on rerun.
                                st.session_state.engine_overrides.pop(
                                    override_key, None
                                )
                                st.session_state[widget_key] = ""
                        if nb_mapped == 0:
                            st.warning(
                                "L'IA n'a proposé aucun mapping exploitable. "
                                "Mappe manuellement ci-dessous."
                            )
                        else:
                            st.success(
                                f"Mapping proposé : {nb_mapped} champ(s) rempli(s). "
                                "Révise-le ci-dessous avant de lancer le moteur."
                            )
                        if notes:
                            st.info(
                                "Notes de l'IA: "
                                + " / ".join(str(n) for n in notes)
                            )
                        if debug:
                            with st.expander("🔍 Debug IA (pour diagnostic)"):
                                st.json({"mapping_proposé": proposed, **debug})
                        st.rerun()

                # Memorize button — Jalon 2.5: build the pattern entry and
                # stash it in session_state. The preview + confirm/cancel
                # section below handles the actual commit to GitHub.
                # Label changes when the file already matches an existing
                # pattern so the user knows it's an update (overwrite by id).
                memorize_label = (
                    "♻️ Mettre à jour ce format"
                    if learned_id
                    else "📋 Mémoriser ce format"
                )
                gh_configured = gh.is_configured()
                memorize_help = (
                    "Construit un pattern à partir du mapping courant. "
                    "Tu vois un aperçu avant que ça parte sur GitHub — "
                    "rien n'est écrit tant que tu n'as pas confirmé."
                    if gh_configured
                    else "Génère un bloc YAML à coller à la main dans "
                         "`src/rules/learned_patterns.yml` (GitHub auto-"
                         "commit non configuré — voir README)."
                )
                if st.button(
                    memorize_label,
                    key=f"engine_memorize_{key}",
                    use_container_width=True,
                    help=memorize_help,
                ):
                    # Effective per-file mapping drawn from the override dict.
                    current_mapping: dict[str, str] = {
                        fld: src_col
                        for (fk, fld), src_col in st.session_state.engine_overrides.items()
                        if fk == key and src_col
                    }
                    if not current_mapping:
                        st.warning(
                            "Aucun champ n'est mappé pour ce fichier — mappe "
                            "d'abord (IA ou manuel), puis clique sur Mémoriser."
                        )
                    else:
                        # Use mapped source columns as the pattern signature:
                        # if these specific headers are present in a future
                        # file, we're confident it's the same format.
                        signature_cols = [c for c in current_mapping.values() if c]
                        author_email = _current_user_email()
                        entry = lp.build_pattern_entry(
                            slug=slug,
                            filename=info["filename"],
                            columns=signature_cols or [str(c) for c in df.columns],
                            loueur_hint=learned_hint or "",
                            column_mapping=current_mapping,
                            author=author_email,
                        )
                        st.session_state[f"engine_pending_pattern_{key}"] = entry
                        # Clear any stale fallback-snippet state from previous runs.
                        st.session_state.pop(f"engine_snippet_{key}", None)
                        st.rerun()
            with c2:
                st.caption(
                    "Colonnes détectées dans le fichier : "
                    + ", ".join(f"`{c}`" for c in list(df.columns)[:10])
                    + (" ..." if len(df.columns) > 10 else "")
                )
                # Delete-pattern button: only shown when this file matched
                # an existing pattern AND GitHub sync is configured. One
                # click = one commit (the deletion is revertible via git
                # history, so we skip the confirmation dance).
                if learned_id and gh.is_configured():
                    if st.button(
                        f"🗑️ Supprimer le pattern `{learned_id}`",
                        key=f"engine_delete_pattern_{key}",
                        help=(
                            "Retire ce pattern de `learned_patterns.yml` sur "
                            "GitHub. Utile si tu as mémorisé par erreur un "
                            "format de test. Réversible via l'historique git."
                        ),
                    ):
                        try:
                            resp = gh.delete_pattern(
                                learned_id,
                                author_email=_current_user_email(),
                            )
                            if resp.get("skipped"):
                                st.warning(
                                    f"Le pattern `{learned_id}` n'existait "
                                    "plus sur GitHub — rien à supprimer."
                                )
                            else:
                                st.success(
                                    f"✅ Pattern `{learned_id}` supprimé "
                                    "sur GitHub. Streamlit va redéployer "
                                    "dans ~1 min — recharge la page après."
                                )
                                # Clear any pending preview for this file
                                # and force a fresh match on next rerun.
                                st.session_state.pop(
                                    f"engine_pending_pattern_{key}", None
                                )
                        except gh.GitHubSyncError as e:
                            st.error(f"❌ {e.user_message}")

            # --- Preview / confirm flow -------------------------------------
            # When the user clicked Memorize we stashed an entry here. Show a
            # YAML preview and require an explicit confirm before committing.
            pending = st.session_state.get(f"engine_pending_pattern_{key}")
            if pending:
                import yaml as _yaml  # local import — yaml is already a dep
                preview_yaml = _yaml.safe_dump(
                    [pending],
                    sort_keys=False,
                    allow_unicode=True,
                    default_flow_style=False,
                )
                st.markdown("##### 🔍 Aperçu du pattern à enregistrer")
                if learned_id and pending.get("id") == learned_id:
                    st.info(
                        f"Ce fichier matchait déjà le pattern `{learned_id}`. "
                        "L'enregistrement va **remplacer** l'entrée existante "
                        "(pas de doublon)."
                    )
                st.code(preview_yaml, language="yaml")

                if gh.is_configured():
                    cc1, cc2 = st.columns([2, 1])
                    with cc1:
                        if st.button(
                            "✅ Confirmer et enregistrer sur GitHub",
                            key=f"engine_confirm_{key}",
                            type="primary",
                            use_container_width=True,
                        ):
                            try:
                                resp = gh.save_pattern(
                                    pending,
                                    author_email=_current_user_email(),
                                )
                                if resp.get("skipped"):
                                    st.info(
                                        "Aucun changement à enregistrer "
                                        "(pattern identique déjà présent)."
                                    )
                                else:
                                    st.success(
                                        f"✅ Pattern `{pending.get('id')}` "
                                        "enregistré sur GitHub. Streamlit va "
                                        "redéployer dans ~1 min — recharge la "
                                        "page après pour que la mémoire soit "
                                        "active."
                                    )
                                st.session_state.pop(
                                    f"engine_pending_pattern_{key}", None
                                )
                            except gh.GitHubSyncError as e:
                                st.error(f"❌ {e.user_message}")
                    with cc2:
                        if st.button(
                            "✖️ Annuler",
                            key=f"engine_cancel_{key}",
                            use_container_width=True,
                        ):
                            st.session_state.pop(
                                f"engine_pending_pattern_{key}", None
                            )
                            st.rerun()
                else:
                    # Fallback: GitHub sync not configured — keep the
                    # copy-paste UX so the feature is still usable.
                    st.warning(
                        "🔑 GitHub auto-commit non configuré (manque "
                        "`GITHUB_TOKEN` et/ou `GITHUB_REPO` dans les secrets "
                        "Streamlit). Copie le bloc ci-dessus sous `patterns:` "
                        "dans `src/rules/learned_patterns.yml` et committe à "
                        "la main."
                    )
                    if st.button(
                        "✖️ Masquer l'aperçu",
                        key=f"engine_cancel_{key}",
                    ):
                        st.session_state.pop(
                            f"engine_pending_pattern_{key}", None
                        )
                        st.rerun()

            st.markdown("---")
            for field_name in MANUAL_MAPPABLE_FIELDS:
                override_key = (key, field_name)
                current = st.session_state.engine_overrides.get(
                    override_key, ""
                )
                if current not in cols:
                    current = ""
                # Layout : selectbox + bouton 🔍 aligné en bas pour peek
                # les valeurs de la colonne source sans quitter la page.
                c_sel, c_peek = st.columns([15, 1], vertical_alignment="bottom")
                with c_sel:
                    picked = st.selectbox(
                        field_name,
                        options=cols,
                        index=cols.index(current),
                        key=f"engine_override_{slug}_{field_name}_{key}",
                    )
                with c_peek:
                    with st.popover(
                        "🔍",
                        help="Aperçu des valeurs de la colonne source",
                        use_container_width=True,
                    ):
                        if picked:
                            _render_column_preview(df, picked)
                        else:
                            st.caption(
                                "Sélectionne d'abord une colonne source pour "
                                "voir un aperçu des valeurs."
                            )
                if picked:
                    st.session_state.engine_overrides[override_key] = picked
                else:
                    st.session_state.engine_overrides.pop(override_key, None)


# ========== Fleet segmentation (Jalon 3.0) ==========

def _fleet_default_file_key(engine_files: dict) -> "str | None":
    """Preselect the most likely « fichier client » file for the dialog.

    Order of preference:
    1. Slug == ``client_file`` (the intended source per spec)
    2. Slug == ``autre_loueur_etat_parc`` (close enough, often used as fallback)
    3. First file in the dict
    """
    if not engine_files:
        return None
    for preferred in ("client_file", "autre_loueur_etat_parc"):
        for k, meta in engine_files.items():
            if meta.get("slug") == preferred:
                return k
    return next(iter(engine_files))


def _fleet_badge_text(mapping) -> str:
    """Short human-readable status text for the badge next to the button."""
    if mapping is None or not mapping.is_active:
        return "Aucune segmentation — 1 seul fichier de sortie."
    names = mapping.fleet_names
    counts = mapping.counts_by_fleet()
    detail = ", ".join(
        f"**{n}** ({counts.get(n, 0)})" for n in names
    )
    return f"**{len(names)} flotte(s)** : {detail}"


@st.dialog("🏢 Segmenter en flottes", width="large")
def _fleet_segmentation_dialog():
    """Popup for configuring the per-import fleet mapping.

    Flow:
    1. Pick the source file (defaults to the detected « fichier client »).
    2. Pick the column that carries the agency / cost-centre code.
    3. Review the unique values + type the display name for each.
       Two raw values mapped to the same display name = they're merged.
       Empty cells surface as « (vide) » so the user never loses rows.
    4. Validate → we build a FleetMapping and stash it in session_state.
    """
    engine_files: dict = st.session_state.engine_files or {}
    if not engine_files:
        st.warning("Charge d'abord au moins un fichier dans l'étape 1.")
        return

    st.caption(
        "Choisis **le fichier** et **la colonne** qui portent le nom "
        "de l'agence / du centre de coût. Tu pourras ensuite renommer / "
        "fusionner / ignorer chaque valeur brute."
    )

    # --- 1. File picker -----------------------------------------------------
    draft = st.session_state.fleet_dialog_draft or {}
    file_keys = list(engine_files.keys())
    labels = [
        f"{engine_files[k].get('filename', k)} · slug={engine_files[k].get('slug', '?')}"
        for k in file_keys
    ]
    default_key = (
        draft.get("file_key")
        or (st.session_state.engine_fleet_mapping.source_file_key
            if st.session_state.engine_fleet_mapping else None)
        or _fleet_default_file_key(engine_files)
    )
    try:
        default_idx = file_keys.index(default_key) if default_key in file_keys else 0
    except ValueError:
        default_idx = 0
    picked_idx = st.selectbox(
        "Fichier source",
        options=list(range(len(file_keys))),
        format_func=lambda i: labels[i],
        index=default_idx,
        key="fleet_dialog_file_idx",
    )
    file_key = file_keys[picked_idx]
    source_df: pd.DataFrame = engine_files[file_key]["df"]

    # --- 2. Column picker ---------------------------------------------------
    all_cols = list(source_df.columns)
    if not all_cols:
        st.error("Ce fichier n'a aucune colonne lisible.")
        return
    suggested = fseg.suggest_agency_columns(source_df)
    default_col = (
        draft.get("column")
        or (st.session_state.engine_fleet_mapping.source_column
            if st.session_state.engine_fleet_mapping
            and st.session_state.engine_fleet_mapping.source_file_key == file_key
            else None)
        or (suggested[0] if suggested else all_cols[0])
    )
    try:
        default_col_idx = all_cols.index(default_col) if default_col in all_cols else 0
    except ValueError:
        default_col_idx = 0
    picked_col = st.selectbox(
        "Colonne « agence »",
        options=all_cols,
        index=default_col_idx,
        key="fleet_dialog_col",
        help="Colonne qui contient le nom / code de l'agence ou du centre de coût.",
    )
    if suggested and picked_col in suggested:
        st.caption(f"💡 Colonne suggérée parmi `{', '.join(suggested)}`.")
    elif suggested:
        st.caption(f"💡 Suggestions si besoin : `{', '.join(suggested)}`.")

    # --- 3. Unique values table --------------------------------------------
    uniques = fseg.unique_values_in_column(source_df, picked_col)
    if not uniques:
        st.warning("Cette colonne est vide ou introuvable.")
        return

    st.markdown("---")
    st.markdown(f"#### 🗂️ Valeurs uniques — {len(uniques)} trouvée(s)")
    st.caption(
        "Tape le **nom de la flotte** à droite. Deux valeurs brutes qui "
        "pointent vers le **même nom** seront **fusionnées**. Laisse un nom "
        "**vide** pour **ignorer** ces lignes (elles n'iront dans aucune flotte)."
    )

    # Build / restore the working raw_to_fleet dict.
    existing = (
        st.session_state.engine_fleet_mapping.raw_to_fleet
        if st.session_state.engine_fleet_mapping
        and st.session_state.engine_fleet_mapping.source_file_key == file_key
        and st.session_state.engine_fleet_mapping.source_column == picked_col
        else {}
    )
    working = dict(draft.get("raw_to_fleet") or existing)

    header = st.columns([4, 1, 5])
    header[0].markdown("**Valeur brute (fichier)**")
    header[1].markdown("**Lignes**")
    header[2].markdown("**Nom de la flotte**")

    for idx, (raw, count) in enumerate(uniques):
        c1, c2, c3 = st.columns([4, 1, 5])
        is_empty = raw == fseg.EMPTY_RAW_KEY
        display_raw = "(vide)" if is_empty else raw
        with c1:
            if is_empty:
                st.markdown(f"_(cellule vide)_")
            else:
                st.markdown(f"`{display_raw}`")
        with c2:
            st.markdown(f"{count}")
        with c3:
            default_value = working.get(raw, "" if is_empty else raw)
            new_val = st.text_input(
                f"flotte_{idx}",
                value=default_value,
                key=f"fleet_dlg_{file_key}_{picked_col}_{idx}",
                label_visibility="collapsed",
                placeholder=("Laisser vide pour ignorer" if is_empty
                             else "ex. Bordeaux"),
            )
            working[raw] = new_val

    # --- Live preview -------------------------------------------------------
    non_empty_targets = [v for v in working.values() if v and v.strip()]
    unique_fleets = sorted(set(v.strip() for v in non_empty_targets))
    st.markdown("---")
    p1, p2, p3 = st.columns(3)
    p1.metric("Flottes distinctes", len(unique_fleets))
    p2.metric("Valeurs mappées", len(non_empty_targets))
    p3.metric("Valeurs ignorées", len(uniques) - len(non_empty_targets))

    # --- Actions ------------------------------------------------------------
    a1, a2, a3 = st.columns([1, 1, 1])
    with a1:
        if st.button("✅ Valider", type="primary", use_container_width=True,
                     key="fleet_dlg_validate"):
            mapping = fseg.build_fleet_mapping(
                source_file_key=file_key,
                source_column=picked_col,
                raw_to_fleet={k: (v or "").strip() for k, v in working.items()},
                source_df=source_df,
            )
            st.session_state.engine_fleet_mapping = mapping
            st.session_state.fleet_dialog_draft = None
            st.rerun()
    with a2:
        if st.button("🧹 Réinitialiser", use_container_width=True,
                     key="fleet_dlg_reset"):
            st.session_state.engine_fleet_mapping = None
            st.session_state.fleet_dialog_draft = None
            st.rerun()
    with a3:
        if st.button("Annuler", use_container_width=True, key="fleet_dlg_cancel"):
            st.session_state.fleet_dialog_draft = None
            st.rerun()

    # Remember the in-flight edits so reopening the dialog doesn't lose them
    # (the dialog closes between reruns unless we stash the draft).
    st.session_state.fleet_dialog_draft = {
        "file_key": file_key,
        "column": picked_col,
        "raw_to_fleet": working,
    }


# ========== Shared engine uploader (Jalon 4.2.2) ==========
def _render_engine_uploader() -> None:
    """Single file_uploader feeding BOTH the Vehicle and Contract tabs.

    Accepts PDF (factures loueurs) + CSV / XLSX (EP, client_file, API
    plaques, …). PDFs are dispatched to ``pdf_parser.parse_factures_to_dataframe``
    with a lessor hint auto-detected from the filename. Tabular files go
    through ``load_tabular`` + ``detect()`` + learned-patterns fallback —
    same logic as before, just unified.

    All parsed files land in ``st.session_state.engine_files`` keyed by
    ``filename::sheet_name`` (xlsx) or ``filename`` (pdf/csv). Any
    previously-computed Vehicle / Contract results are invalidated when
    the upload list changes.
    """
    st.markdown("### 1. Upload des fichiers")
    st.caption(
        "Dépose ici **tous** les fichiers reçus pour le client (PDF factures loueurs, "
        "états de parc, fichier client, API plaques…). Chaque fichier alimente "
        "automatiquement la table à laquelle il correspond — un EP Ayvens ne se "
        "dépose qu'une seule fois et nourrit à la fois Véhicules et Contrats."
    )
    uploaded = st.file_uploader(
        "Fichiers PDF / CSV / XLSX",
        type=["pdf", "csv", "xlsx", "xls", "xlsm"],
        accept_multiple_files=True,
        key="engine_uploader",
    )

    # Only reprocess when the uploaded-file list actually changes. Streamlit
    # returns the uploader's contents on EVERY rerun (including download-
    # button clicks), so without this guard we'd wipe engine_result every
    # time the user downloads an export.
    current_sig = tuple(
        (getattr(f, "name", ""), getattr(f, "size", 0)) for f in (uploaded or [])
    )
    last_sig = st.session_state.get("engine_uploaded_sig")
    if not uploaded or current_sig == last_sig:
        return

    new_files: dict = {}
    # Reset per-file column-mapping overrides up front: file_keys are
    # specific to this upload batch, and stale keys from the previous
    # batch would linger otherwise. We re-populate below for any file
    # that matches a learned pattern with a stored column_mapping.
    st.session_state.engine_overrides = {}
    # Clear any per-file "edit mode" flags from the previous upload batch
    # (Jalon 2.6) and any lingering pending-pattern previews.
    for _stale_key in [
        k for k in list(st.session_state.keys())
        if isinstance(k, str) and (
            k.startswith("engine_edit_recognized_")
            or k.startswith("engine_pending_pattern_")
            or k.startswith("engine_snippet_")
        )
    ]:
        st.session_state.pop(_stale_key, None)

    # Lazy-load the rubriques yaml (for PDF classification). Safe if missing.
    rub_path = Path(__file__).parent / "src" / "rules" / "rubriques_facture.yml"
    whitelist, blacklist = [], []
    if rub_path.exists():
        try:
            import yaml
            rub = yaml.safe_load(rub_path.read_text(encoding="utf-8")) or {}
            whitelist = rub.get("whitelist", [])
            blacklist = rub.get("blacklist", [])
        except Exception as e:
            st.warning(f"Impossible de lire rubriques_facture.yml : {e}")

    _patterns = lp.load_patterns()
    _learned_hits: list[str] = []

    for up in uploaded:
        filename = getattr(up, "name", "uploaded")
        lower = filename.lower()

        # --- PDF branch ---
        if lower.endswith(".pdf"):
            try:
                import tempfile, os as _os
                data = up.read() if hasattr(up, "read") else open(up, "rb").read()
                with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tf:
                    tf.write(data)
                    tmp_path = tf.name
                lessor_hint = None
                if "arval" in lower:
                    lessor_hint = "arval"
                elif "ayvens" in lower or "ald" in lower:
                    lessor_hint = "ayvens"
                df = pdfp.parse_factures_to_dataframe(
                    [tmp_path],
                    whitelist=whitelist,
                    blacklist=blacklist,
                    lessor_hint=lessor_hint,
                )
                try:
                    _os.unlink(tmp_path)
                except Exception:
                    pass
            except Exception as e:
                st.error(f"Échec parsing PDF {filename}: {e}")
                continue
            slug = f"{lessor_hint or 'arval'}_facture_pdf"
            key = filename
            new_files[key] = {
                "df": df,
                "filename": filename,
                "sheet_name": "",
                "slug": slug,
                "detected_type": slug,
                "detected_reason": "PDF parsé",
                "is_pdf": True,
                "learned_match_id": None,
                "learned_match_hint": None,
                "learned_mapping_applied_count": 0,
            }
            continue

        # --- Tabular branch (CSV / XLSX) ---
        try:
            pairs = load_tabular(up)
        except Exception as e:
            st.error(f"Impossible de lire {filename}: {e}")
            continue
        for sheet_name, df in pairs:
            if df.empty or df.shape[1] < 2:
                continue
            key = f"{filename}::{sheet_name}" if sheet_name else filename
            detected = detect(filename, df, sheet_name=sheet_name or None)
            default_slug = DETECTOR_TO_YAML_SLUG.get(detected.source_type, "client_file")
            learned_match = None
            if default_slug == "client_file" and _patterns:
                learned_match = lp.match_pattern(
                    filename, [str(c) for c in df.columns], _patterns
                )
                if learned_match is not None:
                    default_slug = learned_match.slug
                    _learned_hits.append(f"{filename} → `{learned_match.slug}`")

            mapping_applied_count = 0
            if learned_match is not None and learned_match.column_mapping:
                valid_cols = {str(c) for c in df.columns}
                for field_name, src_col in learned_match.column_mapping.items():
                    if src_col and src_col in valid_cols:
                        st.session_state.engine_overrides[(key, field_name)] = src_col
                        mapping_applied_count += 1

            new_files[key] = {
                "df": df,
                "filename": filename,
                "sheet_name": sheet_name,
                "slug": default_slug,
                "detected_type": detected.source_type,
                "detected_reason": detected.reason,
                "is_pdf": False,
                "learned_match_id": learned_match.id if learned_match else None,
                "learned_match_hint": learned_match.loueur_hint if learned_match else None,
                "learned_mapping_applied_count": mapping_applied_count,
            }

    st.session_state.engine_files = new_files
    st.session_state.engine_result = None        # invalidate Vehicle result
    st.session_state.contract_result = None      # invalidate Contract result
    st.session_state.engine_uploaded_sig = current_sig
    st.success(f"{len(new_files)} fichier(s) chargé(s).")
    if _learned_hits:
        st.info(
            "🧠 **Format reconnu depuis la mémoire** — "
            + " · ".join(_learned_hits)
        )


# ========== Shared types-detected section (Jalon 4.2.2) ==========
def _render_engine_types_section(engine_files: dict) -> None:
    """Group loaded files by slug, show per-file selector + preview.

    Works for PDF + XLSX + CSV. The "concat auto" badge is shown when two
    or more files share the same slug (they'll be concatenated with
    `__source_file` lineage before entering the engine).
    """
    st.markdown("### 2. Types détectés — regroupés par source")
    st.caption(
        "Les fichiers sont regroupés par type. Plusieurs fichiers du même type "
        "(ex. deux exports loueurs) sont **concaténés automatiquement** avant "
        "d'entrer dans le moteur (traçabilité via `__source_file`). Corrige "
        "manuellement le type si la détection s'est trompée."
    )

    files_by_slug: dict[str, list[tuple[str, dict]]] = {}
    for key, info in engine_files.items():
        slug = info.get("slug") or "client_file"
        files_by_slug.setdefault(slug, []).append((key, info))

    ordered_slugs = sorted(files_by_slug.keys(), key=lambda s: slug_display(s)[2])
    for slug in ordered_slugs:
        files_in_group = files_by_slug[slug]
        emoji, human_label, _ = slug_display(slug)
        total_rows = sum(len(info["df"]) for _, info in files_in_group)
        n_files = len(files_in_group)

        with st.container(border=True):
            hc1, hc2 = st.columns([5, 2])
            with hc1:
                st.markdown(f"#### {emoji} {human_label}")
                st.caption(
                    f"`{slug}`  ·  {n_files} fichier{'s' if n_files > 1 else ''}  "
                    f"·  {total_rows} ligne{'s' if total_rows > 1 else ''}"
                )
            with hc2:
                if n_files > 1:
                    st.markdown(
                        "<div style='margin-top:0.4rem'><span class='rv-pill' "
                        "style='background:#DBEAFE;color:#1E40AF;border-color:#BFDBFE'>"
                        f"⇢ Concaténation auto · {n_files} fichiers</span></div>",
                        unsafe_allow_html=True,
                    )

            for key, info in files_in_group:
                label = info["filename"] + (
                    f" [{info['sheet_name']}]" if info.get("sheet_name") else ""
                )
                row = st.columns([7, 1, 4])
                with row[0]:
                    st.markdown(f"**{label}**")
                    badges: list[str] = []
                    if info.get("is_pdf"):
                        badges.append(
                            "<span style='background:#FEF3C7;color:#92400E;"
                            "padding:0.08rem 0.5rem;border-radius:999px;"
                            "font-size:0.72rem;font-weight:500;border:1px solid "
                            "#FDE68A'>📄 PDF</span>"
                        )
                    if info.get("learned_match_id"):
                        badges.append(
                            "<span style='background:#F3E8FF;color:#6B21A8;"
                            "padding:0.08rem 0.5rem;border-radius:999px;"
                            "font-size:0.72rem;font-weight:500;border:1px solid "
                            "#E9D5FF'>🧠 mémoire</span>"
                        )
                    badge_html = ("  " + " ".join(badges)) if badges else ""
                    extra = (
                        "PDF parsé"
                        if info.get("is_pdf")
                        else f"détecté `{info.get('detected_type', '?')}` — "
                             f"{info.get('detected_reason', '')}"
                    )
                    st.markdown(
                        f"<span style='color:#64748B;font-size:0.85em'>"
                        f"{len(info['df'])} lignes · {extra}</span>{badge_html}",
                        unsafe_allow_html=True,
                    )
                with row[1]:
                    with st.popover("🔍", help="Aperçu du fichier", use_container_width=True):
                        st.markdown(f"**{label}**")
                        st.caption(
                            f"{len(info['df'])} lignes × {len(info['df'].columns)} colonnes"
                        )
                        st.dataframe(
                            info["df"].head(20),
                            use_container_width=True,
                            hide_index=True,
                        )
                with row[2]:
                    current_slug = info.get("slug", "client_file")
                    if current_slug not in ENGINE_SOURCE_SLUGS:
                        current_slug = "client_file"
                    info["slug"] = st.selectbox(
                        "Type YAML",
                        options=ENGINE_SOURCE_SLUGS,
                        index=ENGINE_SOURCE_SLUGS.index(current_slug),
                        key=f"engine_slug_{key}",
                        label_visibility="collapsed",
                    )


# ========== Shared fleet segmentation controls (Jalon 4.2.2) ==========
def _render_engine_fleet_controls() -> None:
    """Fleet button block — shared between Vehicle + Contract (same mapping
    applies to both tables via plate)."""
    current_fleet_mapping = st.session_state.get("engine_fleet_mapping")
    fc1, fc2 = st.columns([1, 2])
    with fc1:
        if st.button(
            "🏢 Segmenter en flottes",
            use_container_width=True,
            help="Découper l'export en un fichier par agence / centre de coût "
                 "(appliqué aux deux tables Véhicules + Contrats via la plaque).",
            key="fleet_open_dialog",
        ):
            _fleet_segmentation_dialog()
    with fc2:
        st.markdown(
            f"<div style='padding-top:0.55rem;color:#374151;'>"
            f"🏢 {_fleet_badge_text(current_fleet_mapping)}</div>",
            unsafe_allow_html=True,
        )


# ========== Vehicle tab body (Jalon 4.2.2) — mapping + run + results ==========
def _render_vehicle_tab_body(engine_files: dict) -> None:
    """Vehicle-specific workspace: manual mapping for client_file /
    autre_loueur_etat_parc slugs, run button, and results. Assumes
    ``engine_files`` is already populated by the shared uploader."""
    st.caption(
        "Applique les règles de `src/rules/vehicle.yml` et produit le CSV "
        "Vehicle Revio. Les fichiers marqués **`client_file`** ou "
        "**`autre_loueur_etat_parc`** ont besoin d'un mapping manuel (IA + "
        "revue) car leurs colonnes ne sont pas connues à l'avance."
    )

    # --- Column mapping for slugs whose columns aren't baked in YAML ---
    _render_manual_mapping_section(engine_files)

    # --- Run ---
    st.markdown("#### Lancer le moteur Véhicules")

    vehicle_overrides: dict[str, list[str]] = (
        st.session_state.get("rules_overrides", {}).get("vehicle", {})
    )
    vehicle_overrides = {k: v for k, v in vehicle_overrides.items() if v}
    if vehicle_overrides:
        st.info(
            f"🎛️ **{len(vehicle_overrides)} règle(s) de priorité personnalisée(s)** seront appliquées — "
            "cf. *⚙️ Règles d'import* dans le menu de gauche."
        )

    if st.button(
        "▶️ Appliquer les règles Vehicle",
        type="primary",
        use_container_width=True,
        key="engine_run_vehicle_btn",
    ):
        # 1. Split engine_overrides into two structures:
        #    - per_file_overrides: {file_key: {field: source_col}} → used by
        #      merge_engine_sources to rename each file's columns to canonical
        #      names BEFORE concat (handles the case of 2 files of the same
        #      slug with different source headers, e.g. "Plaque d'immatriculation"
        #      vs "Immat.").
        #    - engine_canonical_overrides: {(slug, field): "__map__<field>"}
        #      → tells the rules engine to read the canonical merged column
        #      rather than a specific source header.
        #  Keys of engine_overrides that aren't a known file_key are kept as-is
        #  (belt-and-suspenders: lets the old (slug, field) form keep working
        #  if anything else ever injects it).
        per_file_overrides: dict[str, dict[str, str]] = {}
        engine_canonical_overrides: dict[tuple[str, str], str] = {}
        legacy_slug_overrides: dict[tuple[str, str], str] = {}
        engine_file_keys = set(engine_files.keys())
        for (scope, field_name), src_col in st.session_state.engine_overrides.items():
            if not src_col:
                continue
            if scope in engine_file_keys:
                per_file_overrides.setdefault(scope, {})[field_name] = src_col
                slug_for_file = engine_files[scope].get("slug")
                if slug_for_file:
                    engine_canonical_overrides[(slug_for_file, field_name)] = (
                        f"__map__{field_name}"
                    )
            else:
                # scope is a slug (legacy path) — keep as-is.
                legacy_slug_overrides[(scope, field_name)] = src_col

        # Concat all files sharing the same slug and tag rows with __source_file
        # for downstream traceability. Single-file slugs go through the same
        # code path so the report can always rely on the column existing.
        merged = merge_engine_sources(engine_files, per_file_overrides=per_file_overrides)
        source_dfs: dict = {slug: ms.df for slug, ms in merged.items()}

        # Surface the merges in the UI — if the user dropped two Ayvens files,
        # it's important they SEE that we concatenated them (otherwise they
        # might think the second file got silently dropped).
        merge_notes = [
            f"`{slug}` ← {len(ms.files)} fichiers · {ms.n_rows_before_dedup} lignes"
            for slug, ms in merged.items()
            if len(ms.files) > 1
        ]
        if merge_notes:
            st.info("⇢ **Concaténation auto** — " + "  ·  ".join(merge_notes))

        # Canonical overrides take precedence over legacy slug overrides
        # (canonical columns are guaranteed to exist in the merged df after
        # the per-file rename; legacy source headers may not).
        engine_overrides_for_run = {**legacy_slug_overrides, **engine_canonical_overrides}
        try:
            with st.spinner("Application des règles..."):
                # Load value mappings ONCE per session and reuse the same dict
                # across runs, so AI-proposed pending entries accumulate until
                # the user validates or the page reloads.
                if st.session_state.value_mappings is None:
                    st.session_state.value_mappings = vm.load()
                result = rules_engine.run_vehicle(
                    source_dfs,
                    manual_column_overrides=engine_overrides_for_run,
                    priority_overrides=vehicle_overrides or None,
                    value_mappings=st.session_state.value_mappings,
                )
            # Jalon 2.7 — automatic AI fallback for cells the cache couldn't
            # normalize. Best effort: any error surfaces in the report but
            # doesn't fail the engine result.
            ai_report = None
            if result.unresolved_enums:
                with st.spinner(
                    f"Normalisation IA de {len(result.unresolved_enums)} valeur(s) inconnue(s)..."
                ):
                    ai_report = ain.run_ai_fallback(
                        st.session_state.value_mappings,
                        result.unresolved_enums,
                        df=result.df,
                        orphan_df=result.orphan_df,
                        user_email=st.session_state.get("current_user") or None,
                        value_mapping_hits=result.value_mapping_hits,
                    )
                # Best-effort GitHub auto-commit of the newly proposed
                # pending entries — only if anything was actually upserted
                # and GitHub is configured.
                if ai_report and ai_report.total_proposed > 0 and gh.is_configured():
                    try:
                        yaml_text = vm.dump_yaml(st.session_state.value_mappings)
                        gh.save_value_mappings_yaml(
                            yaml_text,
                            commit_message=(
                                f"value_mappings: +{ai_report.total_proposed} "
                                f"pending via AI fallback"
                            ),
                            author_email=st.session_state.get("current_user") or None,
                        )
                    except Exception as _gh_err:
                        # Silent for now; the « 🧭 Normalisation » page will
                        # show the un-committed state.
                        pass
            st.session_state.engine_result = result
            st.session_state.ai_fallback_report = ai_report
        except Exception as e:
            st.error(f"Erreur moteur: {e}")
            st.session_state.engine_result = None
            st.session_state.ai_fallback_report = None

    # --- Results ---
    result = st.session_state.engine_result
    if result is None:
        return

    st.markdown("#### Résultat Véhicules")
    df = result.df
    col_a, col_b, col_c = st.columns(3)
    col_a.metric("Véhicules produits", len(df))
    col_b.metric("Anomalies détectées", len(result.conflicts_by_cell or {}))
    orphan_count = len(result.orphan_df) if result.orphan_df is not None else 0
    col_c.metric("Plaques orphelines", orphan_count)
    st.dataframe(df, use_container_width=True)

    if result.issues:
        st.markdown(f"#### ⚠️ {len(result.issues)} alerte(s) globale(s)")
        issues_df = pd.DataFrame([
            {"plaque": i.plate, "champ": i.field, "source": i.source, "avertissement": i.warning}
            for i in result.issues
        ])
        st.dataframe(issues_df, use_container_width=True, hide_index=True)

    if result.orphan_df is not None and not result.orphan_df.empty:
        st.markdown(f"#### 👻 {len(result.orphan_df)} plaque(s) orpheline(s)")
        st.caption(
            "Ces plaques sont présentes dans un fichier loueur mais absentes du fichier "
            "client. Elles sont exclues du parc final mais listées dans l'onglet "
            "`plaques_orphelines` du rapport Excel."
        )
        st.dataframe(result.orphan_df, use_container_width=True)

    # --- Normalisations (Jalon 2.7) ----------------------------------
    _render_normalizations_report(result, st.session_state.ai_fallback_report)

    if not result.conflicts_by_cell and not result.issues and orphan_count == 0:
        st.success("Aucune anomalie — toutes les sources se sont bien alignées ✅")


# ========== Unified download (shared between Vehicle + Contract tabs) ==========
def _render_unified_download() -> None:
    """Single download button that packages everything the engines produced.

    Pulls both Vehicle and Contract results from session state. The zip is
    always valid — if only one base ran, only that one lands in the zip.
    """
    v_result = st.session_state.get("engine_result")
    c_result = st.session_state.get("contract_result")
    if v_result is None and c_result is None:
        return

    st.markdown("### 📦 Export — pack unifié")
    client_name = st.session_state.client_name or "client"
    fleet_mapping = st.session_state.get("engine_fleet_mapping")

    # Summary line.
    bits: list[str] = []
    if v_result is not None and v_result.df is not None and not v_result.df.empty:
        bits.append(f"{len(v_result.df)} véhicule(s)")
    if c_result is not None and c_result.df is not None and not c_result.df.empty:
        bits.append(f"{len(c_result.df)} contrat(s)")
    if fleet_mapping is not None and fleet_mapping.is_active:
        bits.append(f"segmenté en {len(fleet_mapping.fleet_names)} flotte(s)")
    st.caption("📦 Le pack contiendra : " + " · ".join(bits) + " + rapports + classeur maître.")

    # Vehicle report (existing).
    vehicle_report_bytes = None
    if v_result is not None:
        try:
            vehicle_report_bytes = build_report_xlsx(v_result, client_name=client_name)
        except Exception as e:
            st.warning(f"Rapport Excel véhicules indisponible : {e}")

    # Contract df + errors df (issues + conflicts → one sheet each).
    contract_df = c_result.df if c_result is not None else None
    contract_orphan_df = c_result.orphan_df if c_result is not None else None
    contract_errors_bytes = None
    if c_result is not None:
        try:
            contract_errors_bytes = _build_contract_errors_xlsx(c_result, client_name=client_name)
        except Exception as e:
            st.warning(f"Rapport Excel contrats indisponible : {e}")

    # Lineage sidecars (parquet if pyarrow available, jsonl fallback).
    lineage_bytes_by_name: dict[str, bytes] = {}
    for label, store in (
        ("vehicle", getattr(v_result, "lineage", None) if v_result else None),
        ("contract", getattr(c_result, "lineage", None) if c_result else None),
    ):
        if store is None or len(store) == 0:
            continue
        try:
            import io as _io
            df = store.to_dataframe()
            buf = _io.BytesIO()
            try:
                df.to_parquet(buf, engine="pyarrow", index=False)
                lineage_bytes_by_name[f"_lineage/{label}.parquet"] = buf.getvalue()
            except Exception:
                # Fallback to JSONL if pyarrow missing
                buf = _io.StringIO()
                df.to_json(buf, orient="records", lines=True, force_ascii=False)
                lineage_bytes_by_name[f"_lineage/{label}.jsonl"] = buf.getvalue().encode("utf-8")
        except Exception:
            pass  # lineage is additive; never block the download on it

    try:
        zip_bytes, zip_name = zw.build_output_zip(
            client_name=client_name,
            vehicle_df=v_result.df if v_result is not None else None,
            vehicle_fleet_mapping=fleet_mapping,
            report_xlsx_bytes=vehicle_report_bytes,
            contract_df=contract_df,
            contract_orphan_df=contract_orphan_df,
            contract_errors_xlsx_bytes=contract_errors_bytes,
            contract_fleet_mapping=fleet_mapping,  # shared fleet mapping by plate
            extra_files=lineage_bytes_by_name,
        )
        st.download_button(
            "📦 Télécharger le pack d'import (.zip)",
            data=zip_bytes,
            file_name=zip_name,
            mime="application/zip",
            type="primary",
            use_container_width=True,
        )
    except Exception as e:
        st.error(f"Impossible de construire le zip : {e}")


def _build_contract_errors_xlsx(result, *, client_name: str) -> bytes:
    """Produce a minimal contracts_errors.xlsx — 1 sheet per type (issues, conflicts).

    Each row = 1 anomaly, with enough context to resolve. Cell-by-cell format
    (R11 of the Contract spec). No inter-table duplication (isHT conflicts
    already surface in the Vehicle errors file).
    """
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: global issues
    ws = wb.create_sheet("issues")
    ws.append(["plate", "number", "field", "source", "avertissement"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="1F2937")
    for i in result.issues or []:
        ws.append([
            getattr(i, "plate", ""),
            getattr(i, "number", ""),
            getattr(i, "field", ""),
            getattr(i, "source", ""),
            getattr(i, "warning", ""),
        ])

    # Sheet 2: per-cell conflicts (tolerance-aware)
    ws2 = wb.create_sheet("conflits_cellule")
    ws2.append(["plate", "number", "field", "valeur_retenue", "conflits"])
    for c in ws2[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="1F2937")
    for (key, field), conflicts in (result.conflicts_by_cell or {}).items():
        plate, _, number = str(key).partition("|")
        winner_src = (result.source_by_cell or {}).get((key, field))
        ws2.append([
            plate, number, field, str(winner_src),
            " | ".join(f"{c.get('source')}={c.get('value')} ({c.get('reason', '')})" for c in conflicts),
        ])

    # Sheet 3: orphelins (contrats dans factures/EP loueur mais absents client_file)
    ws3 = wb.create_sheet("orphelins")
    if result.orphan_df is not None and not result.orphan_df.empty:
        cols = list(result.orphan_df.columns)
        ws3.append(cols)
        for c in ws3[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", start_color="1F2937")
        for _, row in result.orphan_df.iterrows():
            ws3.append([row.get(c, "") for c in cols])
    else:
        ws3.append(["Aucun contrat orphelin détecté."])

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ========== Engine page — unified uploader + table-centric tabs (Jalon 4.2.2) ==========
def render_engine_page():
    """Top-level Moteur page (Jalon 4.2.2).

    Layout:
    1. Shared file_uploader (PDF + CSV + XLSX) — one deposit, both tables.
    2. Shared "Types détectés — regroupés par source" section.
    3. Shared fleet segmentation controls.
    4. Two tabs — 🚗 Véhicules and 📄 Contrats — each being its own mapping
       + run + results workspace (no upload inside).
    5. Shared download button producing the unified zip.

    Intent (Augustin, 2026-04-22): "Un seul endroit pour déposer tous les
    docs. Plusieurs onglets pour définir, si besoin, les mappings. Un seul
    bouton téléchargement en bas."
    """
    st.title("🧪 Import — Moteur de règles")

    _render_engine_uploader()
    engine_files = st.session_state.get("engine_files") or {}
    if not engine_files:
        st.info(
            "Dépose un ou plusieurs fichiers ci-dessus pour démarrer. "
            "Un même fichier (ex. EP Ayvens) alimente automatiquement les tables "
            "Véhicules **et** Contrats — pas besoin de l'envoyer deux fois."
        )
        return

    _render_engine_types_section(engine_files)
    _render_engine_fleet_controls()

    tab_v, tab_c = st.tabs(["🚗 Véhicules", "📄 Contrats"])
    with tab_v:
        _render_vehicle_tab_body(engine_files)
    with tab_c:
        _render_contract_tab_body(engine_files)

    st.markdown("---")
    _render_unified_download()


# ========== Contract tab body (Jalon 4.2.2) — run + results + unknown cols ==========
def _render_contract_tab_body(engine_files: dict) -> None:
    """Contract-specific workspace: run button, results, and the unknown-
    columns UI. Assumes ``engine_files`` is already populated by the shared
    uploader — files whose slugs aren't declared in ``contract.yml`` are
    silently ignored by the Contract engine."""
    st.caption(
        "Applique les règles de `src/rules/contract.yml` et produit un CSV "
        "contrats prêt pour Revio, clé primaire **(plaque, numéro de contrat)**. "
        "Les factures PDF déposées en haut sont automatiquement parsées et "
        "reconnues ici."
    )
    if not rules_io.list_available_tables():
        st.warning("contract.yml introuvable — le moteur Contract n'est pas encore configuré.")
        return

    # --- Run ---
    st.markdown("#### Lancer le moteur Contrats")

    # Carry VP info from the Vehicle result (if any) so isHT post-pass can
    # reference the canonical VP classification computed earlier.
    vehicle_vp_by_plate = _extract_vp_from_vehicle_result(st.session_state.get("engine_result"))
    if vehicle_vp_by_plate:
        st.caption(
            f"ℹ️ VP hérité du moteur Véhicules : {sum(1 for v in vehicle_vp_by_plate.values() if v)} "
            f"VP / {sum(1 for v in vehicle_vp_by_plate.values() if not v)} non-VP."
        )

    if st.button("▶️ Appliquer les règles Contract", type="primary",
                 use_container_width=True, key="engine_run_contract_btn"):
        # Build source_dfs keyed by slug (concat files of same slug). The
        # Contract engine ignores slugs it doesn't understand, so it's safe
        # to pass the full shared engine_files here.
        source_dfs: dict = {}
        for key, info in engine_files.items():
            slug = info.get("slug")
            if not slug:
                continue
            df = info.get("df")
            if df is None or df.empty:
                continue
            if slug in source_dfs:
                source_dfs[slug] = pd.concat([source_dfs[slug], df], ignore_index=True, sort=False)
            else:
                source_dfs[slug] = df.copy()

        # Apply user-resolved unknown columns (register_learned_column + overrides)
        overrides: dict[tuple[str, str], str] = {}
        for (src_slug, field_name), col in st.session_state.contract_unknown_resolved.items():
            overrides[(src_slug, field_name)] = col

        try:
            with st.spinner("Application des règles Contract..."):
                result = ce.run_contract(
                    source_dfs=source_dfs,
                    manual_column_overrides=overrides or None,
                    vehicle_vp_by_plate=vehicle_vp_by_plate,
                )
            st.session_state.contract_result = result
        except Exception as e:
            st.error(f"Erreur moteur Contract : {e}")
            st.session_state.contract_result = None

    result = st.session_state.get("contract_result")
    if result is None:
        return

    # --- Results ---
    st.markdown("#### Résultat Contrats")
    df = result.df
    col_a, col_b, col_c, col_d = st.columns(4)
    col_a.metric("Contrats produits", len(df))
    col_b.metric("Conflits détectés", len(result.conflicts_by_cell or {}))
    col_c.metric("Contrats orphelins", len(result.orphan_df) if result.orphan_df is not None else 0)
    col_d.metric("Colonnes non identifiées", len(result.unknown_column_requests or []))

    st.dataframe(df, use_container_width=True)

    if result.issues:
        st.markdown(f"##### ⚠️ {len(result.issues)} alerte(s)")
        issues_rows = [
            {"plaque": i.plate, "numéro": getattr(i, "number", ""),
             "champ": i.field, "source": i.source, "avertissement": i.warning}
            for i in result.issues
        ]
        st.dataframe(pd.DataFrame(issues_rows), use_container_width=True, hide_index=True)

    if result.orphan_df is not None and not result.orphan_df.empty:
        st.markdown(f"##### 👻 {len(result.orphan_df)} contrat(s) orphelin(s)")
        st.caption(
            "Ces contrats sont présents dans un fichier loueur mais absents du "
            "fichier client. Ils seront listés dans l'onglet `orphelins` du "
            "rapport d'erreurs Contract."
        )
        st.dataframe(result.orphan_df, use_container_width=True)

    # --- Unknown columns UI (Jalon 4.1.7) ---
    if result.unknown_column_requests:
        _render_contract_unknown_columns_ui(result.unknown_column_requests, engine_files)


def _extract_vp_from_vehicle_result(vehicle_result) -> dict[str, bool]:
    """Inspect Vehicle engine output and return {plate_norm: is_vp} for the
    Contract engine's isHT post-pass. Returns empty dict if no result."""
    if vehicle_result is None or vehicle_result.df is None:
        return {}
    df = vehicle_result.df
    if "usage" not in df.columns:
        return {}
    out: dict[str, bool] = {}
    plate_col = "registrationPlate" if "registrationPlate" in df.columns else None
    if plate_col is None:
        return {}
    from src.normalizers import plate_for_matching
    for _, row in df.iterrows():
        p = plate_for_matching(row.get(plate_col))
        if not p:
            continue
        usage = str(row.get("usage") or "").strip().lower()
        # private → VP ; utility/service → non-VP
        if usage == "private":
            out[p] = True
        elif usage in ("utility", "service"):
            out[p] = False
    return out


def _render_contract_unknown_columns_ui(requests: list, engine_files: dict) -> None:
    """Ask the user to resolve the mandatory fields the engine couldn't fill.

    On validation we persist via `unknown_columns.register_learned_column`
    so the next run on a similar file auto-resolves. The resolved picks are
    also mirrored in session state so the user can re-run immediately.

    `engine_files` is the unified Jalon 4.2.2 dict (same shape as before,
    shared between Vehicle + Contract tabs).
    """
    st.markdown("##### 🧩 Colonnes non identifiées")
    st.caption(
        "Pour ces champs obligatoires, le moteur n'a pas pu repérer la colonne source. "
        "Choisis la bonne colonne une fois — l'app la retient pour les prochains imports."
    )
    patterns_path = Path(__file__).parent / "src" / "rules" / "learned_patterns.yml"
    user_email = st.session_state.get("current_user") or _current_user_email()

    # Map slug → list of available columns (picked from the corresponding file df)
    cols_by_slug: dict[str, list[str]] = {}
    df_by_slug: dict[str, pd.DataFrame] = {}
    for key, info in engine_files.items():
        slug = info.get("slug")
        if not slug:
            continue
        df = info.get("df")
        if df is None or df.empty:
            continue
        cols_by_slug[slug] = [str(c) for c in df.columns]
        df_by_slug[slug] = df

    for req in requests:
        field = req.get("field") if isinstance(req, dict) else req.field
        plate = req.get("plate") if isinstance(req, dict) else req.plate
        number = req.get("number") if isinstance(req, dict) else req.number
        candidates = req.get("candidate_sources") if isinstance(req, dict) else req.candidate_sources
        hint = req.get("hint") if isinstance(req, dict) else req.hint

        available_slugs = [s for s in candidates if s in cols_by_slug] or (["client_file"] if "client_file" in cols_by_slug else [])
        if not available_slugs:
            st.warning(f"Aucune source chargée ne peut porter `{field}` ({plate} / {number}).")
            continue

        with st.container(border=True):
            st.markdown(f"**Champ manquant : `{field}`** — contrat {plate} / {number}")
            if hint:
                st.caption(f"💡 {hint}")
            pick_slug = st.selectbox(
                "Source",
                options=available_slugs,
                key=f"unk_slug_{field}_{plate}_{number}",
            )
            available_cols = [""] + cols_by_slug.get(pick_slug, [])
            prev = st.session_state.contract_unknown_resolved.get((pick_slug, field), "")
            default_idx = available_cols.index(prev) if prev in available_cols else 0
            pick_col = st.selectbox(
                "Colonne",
                options=available_cols,
                index=default_idx,
                key=f"unk_col_{field}_{plate}_{number}",
            )
            c1, c2 = st.columns(2)
            with c1:
                if st.button("💾 Mémoriser + appliquer",
                             key=f"unk_save_{field}_{plate}_{number}",
                             use_container_width=True,
                             type="primary",
                             disabled=not pick_col):
                    try:
                        unkcol.register_learned_column(
                            patterns_path,
                            table="contract",
                            source_slug=pick_slug,
                            field_name=field,
                            column=pick_col,
                            source_df=df_by_slug.get(pick_slug),
                            learned_by=user_email,
                        )
                        st.session_state.contract_unknown_resolved[(pick_slug, field)] = pick_col
                        st.success(f"✓ `{field}` ← `{pick_col}` mémorisé. Relance le moteur pour appliquer.")
                        # Best-effort commit on GitHub (same pattern as value_mappings)
                        try:
                            if gh.is_configured():
                                yml_text = patterns_path.read_text(encoding="utf-8")
                                gh.save_learned_patterns_yaml(
                                    yml_text,
                                    commit_message=f"learned_patterns: contract.{pick_slug}.{field} = {pick_col}",
                                    author_email=user_email,
                                )
                        except Exception:
                            pass  # GitHub commit is best-effort
                    except Exception as e:
                        st.error(f"Échec sauvegarde : {e}")
            with c2:
                if (pick_slug, field) in st.session_state.contract_unknown_resolved:
                    if st.button("↺ Oublier ce mapping",
                                 key=f"unk_forget_{field}_{plate}_{number}",
                                 use_container_width=True):
                        st.session_state.contract_unknown_resolved.pop((pick_slug, field), None)
                        st.rerun()


# ========== Rules editor page ==========
def render_rules_page():
    st.header("⚙️ Règles d'import")
    st.caption(
        "Configure ici, champ par champ, quelle source gagne sur les autres quand "
        "plusieurs fichiers donnent une valeur différente pour la même information. "
        "Ces règles pilotent le moteur de l'onglet *Import — Moteur de règles*."
    )

    # --- Banner: session scope ---
    st.warning(
        "🔒 **Portée session uniquement** — Les modifications faites ici s'appliquent "
        "à l'onboarding en cours. Elles ne touchent pas les règles par défaut et seront "
        "perdues à la fermeture de l'app. Utilise *🔄 Réinitialiser tout* pour revenir aux "
        "valeurs par défaut à tout moment."
    )

    overrides_all = st.session_state.setdefault("rules_overrides", {})
    nb_modified = rules_io.count_active_overrides(overrides_all)

    # --- Summary bar ---
    sc1, sc2 = st.columns([4, 1])
    with sc1:
        if nb_modified > 0:
            st.info(f"✎ **{nb_modified} champ(s) personnalisé(s)** dans cette session.")
        else:
            st.success("✅ Aucune modification — les règles par défaut s'appliquent.")
    with sc2:
        if nb_modified > 0:
            if st.button("🔄 Réinitialiser tout", use_container_width=True, key="rules_reset_all"):
                st.session_state.rules_overrides = {}
                st.rerun()

    # --- Legend ---
    with st.expander("ℹ️ Comment lire cette page ?", expanded=False):
        st.markdown(
            "**Numéro à gauche** : priorité de la source pour ce champ — "
            "plus le chiffre est petit, plus elle passe en premier.\n\n"
            "**Ex-æquo** : plusieurs sources peuvent avoir la même priorité "
            "(ex : Ayvens, Arval, autre loueur à `2` pour `brand`). En pratique "
            "une même plaque n'est présente que chez un seul loueur → c'est "
            "celui-là qui gagne sur sa plaque, sans conflit.\n\n"
            "**🟢 Source chargée** pour cet import — la règle va s'appliquer.  \n"
            "**⚪ Source non chargée** — la règle est là mais inerte tant qu'un "
            "fichier correspondant n'est pas déposé dans le moteur.\n\n"
            "**⬆ / ⬇** : remonter / descendre une source. Réorganiser **casse "
            "les ex-æquo** et fige un ordre strict pour ce champ. Utilise *↺ "
            "Réinitialiser* pour retrouver les priorités par défaut."
        )

    st.markdown("---")

    # --- Table selection (tabs) ---
    tables = rules_io.list_available_tables()
    tab_labels = [
        meta["label"] + ("" if meta["available"] else " — bientôt")
        for _, meta in tables
    ]
    tabs = st.tabs(tab_labels)
    for i, (slug, meta) in enumerate(tables):
        with tabs[i]:
            if meta["available"]:
                _render_table_rules(slug)
            else:
                st.info(
                    f"Les règles **{meta['label']}** seront disponibles dans une prochaine "
                    "version. Aujourd'hui, seule la table Véhicules est câblée au moteur."
                )


# ========== Normalisation page — value-level mapping dictionary ==========
def render_normalization_page():
    """Editor for ``src/rules/value_mappings.yml``.

    Shows one expander per enum-typed field (Usage, Motorisation, booleans, ...),
    each listing its raw → canonical entries. For every row the user can:
    - ✅ validate a pending IA proposition (status pending → validated),
    - ✏️ change the canonical target (dropdown in allowed_values),
    - 🗑️ delete the entry,
    - ➕ add a brand-new mapping via the per-field form.

    All edits mutate ``st.session_state.value_mappings`` in place. The
    « 💾 Sauvegarder sur GitHub » button serializes the dict via
    ``vm.dump_yaml`` and pushes it through ``gh.save_value_mappings_yaml``.
    """
    st.header("🧭 Normalisation des valeurs")
    st.caption(
        "Dictionnaire qui transforme les valeurs brutes des fichiers clients "
        "(VP/VU, Essence, Hybride-diesel, Oui/Non…) en valeurs canoniques Revio "
        "(private/utility, diesel/gas/hybrid/electric, TRUE/FALSE…). "
        "Les propositions **IA en attente** sont appliquées tout de suite mais "
        "attendent ta validation avant d'être réutilisées sans doute."
    )

    # Lazy-load the dict from disk on first visit.
    if st.session_state.get("value_mappings") is None:
        st.session_state.value_mappings = vm.load()
    mappings: dict = st.session_state.value_mappings

    # --- Collect every enum-typed field defined in SCHEMAS ---
    # Used to (a) show empty buckets when a field has no entry yet,
    # (b) bound the « target » selectors to valid allowed_values.
    enum_fields = vm.iter_enum_fields(SCHEMAS)  # list of (schema, field, allowed)
    allowed_by_fkey: dict[str, list[str]] = {
        vm.field_key(sc, fn): vals for (sc, fn, vals) in enum_fields
    }

    # --- Global stats bar ---
    stats = vm.stats_by_field(mappings)
    total_validated = sum(s.get(vm.STATUS_VALIDATED, 0) for s in stats.values())
    total_pending = sum(s.get(vm.STATUS_PENDING, 0) for s in stats.values())
    total_fields = len(allowed_by_fkey)

    c1, c2, c3, c4 = st.columns([1, 1, 1, 2])
    c1.metric("✅ Validées", total_validated)
    c2.metric("🤖 En attente", total_pending)
    c3.metric("Champs enum", total_fields)
    with c4:
        st.write("")  # vertical align with metrics
        if st.button(
            "💾 Sauvegarder sur GitHub",
            type="primary",
            use_container_width=True,
            disabled=not gh.is_configured(),
            help=(
                "Commit le dictionnaire dans le repo GitHub pour qu'il soit "
                "partagé avec les autres utilisateurs."
                if gh.is_configured()
                else "GitHub non configuré — les secrets TOKEN/REPO sont absents."
            ),
        ):
            try:
                yaml_text = vm.dump_yaml(mappings)
                gh.save_value_mappings_yaml(
                    yaml_text,
                    commit_message=(
                        f"value_mappings: manual edit — {total_validated} "
                        f"validées / {total_pending} en attente"
                    ),
                    author_email=_current_user_email(),
                )
                st.success("Dictionnaire sauvegardé sur GitHub. ✅")
            except Exception as e:
                st.error(f"Échec de la sauvegarde GitHub : {e}")

    if total_pending > 0:
        st.info(
            f"🤖 **{total_pending} proposition(s) IA en attente de validation.** "
            "Déroule les champs concernés ci-dessous pour valider ou corriger."
        )

    st.markdown("---")

    # --- Helper: friendlier display labels for common enum fields ---
    friendly_labels = {
        "vehicle.usage": "🚗 Véhicule — usage (private / utility / service)",
        "vehicle.motorisation": "⛽ Véhicule — motorisation",
        "driver.civility": "🧑 Conducteur — civilité (M / F)",
        "driver.seniority": "🧑 Conducteur — séniorité",
        "driver.professionalStatus": "🧑 Conducteur — statut (internal / external)",
        "contract.isHT": "📄 Contrat — HT (TRUE / FALSE)",
        "contract.maintenanceNetwork": "📄 Contrat — réseau maintenance",
        "contract.tiresType": "📄 Contrat — type de pneus",
        "contract.tiresNetwork": "📄 Contrat — réseau pneus",
        "asset.kind": "💳 Asset — type (fuel_card / toll_tag)",
    }

    # --- Sort: fields with pending items first, then by label ---
    def _sort_key(fkey: str) -> tuple[int, str]:
        pending = stats.get(fkey, {}).get(vm.STATUS_PENDING, 0)
        label = friendly_labels.get(fkey, fkey)
        return (0 if pending > 0 else 1, label.lower())

    # Union of all known enum fields and fields present in the dict
    # (so booleans like civilLiabilityEnabled show up even when empty).
    all_fkeys = sorted(
        set(allowed_by_fkey.keys()) | set(mappings.keys()),
        key=_sort_key,
    )

    for fkey in all_fkeys:
        schema, field_name = vm.parse_field_key(fkey)
        bucket = mappings.get(fkey, {})
        allowed = allowed_by_fkey.get(fkey, [])
        counts = stats.get(fkey, {})
        n_val = counts.get(vm.STATUS_VALIDATED, 0)
        n_pen = counts.get(vm.STATUS_PENDING, 0)
        label = friendly_labels.get(fkey, f"`{fkey}`")
        badge = []
        if n_val:
            badge.append(f"✅ {n_val}")
        if n_pen:
            badge.append(f"🤖 {n_pen}")
        if not badge:
            badge.append("(vide)")
        title = f"{label} — {' · '.join(badge)}"
        with st.expander(title, expanded=(n_pen > 0)):
            _render_field_mappings(
                mappings=mappings,
                schema=schema,
                field_name=field_name,
                bucket=bucket,
                allowed=allowed,
            )


def _render_field_mappings(
    mappings: dict,
    schema: str,
    field_name: str,
    bucket: dict,
    allowed: list[str],
) -> None:
    """Inline editor for a single (schema, field) entry list.

    Each row shows raw → target + status + source, with per-row buttons
    (validate, delete). Target is edited inline with a selectbox bounded
    to ``allowed``. Below the list, a small form lets the user add a new
    manual mapping.
    """
    if not allowed:
        st.caption(
            f"⚠️ Ce champ n'a pas de `allowed_values` déclarés dans SCHEMAS — "
            "ajoute-le à `src/schemas.py` pour activer l'éditeur."
        )
        return

    # --- Allowed values reminder (one-liner) ---
    st.caption("Valeurs Revio autorisées : " + ", ".join(f"`{v}`" for v in allowed))

    if not bucket:
        st.info("Aucune entrée pour ce champ — utilise le formulaire ci-dessous.")
    else:
        # Sort: pending first, then by target then raw.
        def _row_sort(kv):
            key, vm_entry = kv
            return (0 if vm_entry.status == vm.STATUS_PENDING else 1,
                    vm_entry.target, vm_entry.raw.lower())

        rows = sorted(bucket.items(), key=_row_sort)
        # Header line
        h1, h2, h3, h4, h5, h6 = st.columns([3, 2, 1.4, 1.4, 0.8, 0.8])
        h1.markdown("**Valeur brute**")
        h2.markdown("**→ Revio**")
        h3.markdown("**Statut**")
        h4.markdown("**Source**")
        h5.markdown("**✅**")
        h6.markdown("**🗑**")

        for norm_key, entry in rows:
            c1, c2, c3, c4, c5, c6 = st.columns([3, 2, 1.4, 1.4, 0.8, 0.8])
            with c1:
                raw_display = entry.raw if entry.raw.strip() else "(vide)"
                note = f" — _{entry.note}_" if entry.note else ""
                st.markdown(f"`{raw_display}`{note}")
            with c2:
                new_target = st.selectbox(
                    "target",
                    options=allowed,
                    index=allowed.index(entry.target) if entry.target in allowed else 0,
                    key=f"norm_target_{schema}_{field_name}_{norm_key}",
                    label_visibility="collapsed",
                )
                if new_target != entry.target:
                    vm.upsert(
                        mappings, schema, field_name, entry.raw, new_target,
                        status=vm.STATUS_VALIDATED,
                        source=vm.SOURCE_MANUAL,
                        user=_current_user_email(),
                        note=entry.note,
                    )
                    st.rerun()
            with c3:
                if entry.status == vm.STATUS_VALIDATED:
                    st.success("validée")
                else:
                    st.warning("en attente")
            with c4:
                st.caption(entry.source)
            with c5:
                if entry.status == vm.STATUS_PENDING:
                    if st.button(
                        "✅",
                        key=f"norm_validate_{schema}_{field_name}_{norm_key}",
                        help="Valider cette proposition IA",
                        use_container_width=True,
                    ):
                        vm.validate_entry(
                            mappings, schema, field_name, norm_key,
                            user=_current_user_email(),
                        )
                        st.rerun()
                else:
                    st.caption(" ")
            with c6:
                if st.button(
                    "🗑",
                    key=f"norm_delete_{schema}_{field_name}_{norm_key}",
                    help="Supprimer cette correspondance",
                    use_container_width=True,
                ):
                    vm.delete_entry(mappings, schema, field_name, norm_key)
                    st.rerun()

    # --- Add new mapping form ---
    st.markdown("")
    with st.form(key=f"norm_add_{schema}_{field_name}", clear_on_submit=True):
        f1, f2, f3 = st.columns([3, 2, 1])
        with f1:
            new_raw = st.text_input(
                "Nouvelle valeur brute",
                key=f"norm_new_raw_{schema}_{field_name}",
                placeholder="ex. Hybride-Essence",
            )
        with f2:
            new_target = st.selectbox(
                "Valeur Revio",
                options=allowed,
                key=f"norm_new_target_{schema}_{field_name}",
            )
        with f3:
            st.write("")  # align with inputs
            submitted = st.form_submit_button(
                "➕ Ajouter", use_container_width=True, type="primary",
            )
        if submitted:
            if not new_raw or not new_raw.strip():
                st.warning("La valeur brute ne peut pas être vide.")
            else:
                try:
                    vm.upsert(
                        mappings, schema, field_name,
                        new_raw.strip(), new_target,
                        status=vm.STATUS_VALIDATED,
                        source=vm.SOURCE_MANUAL,
                        user=_current_user_email(),
                    )
                    st.success(
                        f"Ajouté : `{new_raw.strip()}` → `{new_target}` ✅"
                    )
                    st.rerun()
                except ValueError as e:
                    st.error(str(e))


def _render_table_rules(table_slug: str):
    """Render the priority editor for a single table (vehicle / contract / ...)."""
    try:
        rules_yaml = rules_io.load_rules_yaml(table_slug)
    except (FileNotFoundError, KeyError) as e:
        st.error(f"Impossible de charger les règles : {e}")
        return

    fields_spec = rules_yaml.get("fields", {})
    if not fields_spec:
        st.info("Aucun champ déclaré dans ce fichier de règles.")
        return

    overrides_table = st.session_state.setdefault("rules_overrides", {}).setdefault(
        table_slug, {}
    )

    # Which source slugs are currently uploaded in the engine? Used to paint
    # the "effective" source badges so the user sees at a glance which priority
    # levels will actually fire for THIS import.
    uploaded_slugs: set[str] = {
        info.get("slug")
        for info in st.session_state.get("engine_files", {}).values()
        if info.get("slug")
    }

    # --- Search bar ---
    sb1, sb2 = st.columns([4, 1])
    with sb1:
        query = st.text_input(
            "🔎 Rechercher un champ",
            value="",
            placeholder="ex. usage, motorisation, co2, brand…",
            key=f"rules_search_{table_slug}",
            label_visibility="collapsed",
        ).strip().lower()
    with sb2:
        show_only_modified = st.toggle(
            "Modifiés uniquement",
            value=False,
            key=f"rules_only_modified_{table_slug}",
        )

    # --- Group by category ---
    categories = rules_io.categorize_fields(table_slug, fields_spec)
    cat_tabs = st.tabs([c[0] for c in categories])

    for i, (cat_label, field_names) in enumerate(categories):
        with cat_tabs[i]:
            # Filter per search / modified toggle
            filtered = []
            for fname in field_names:
                if fname not in fields_spec:
                    continue
                if query and query not in fname.lower() and query not in (
                    fields_spec[fname].get("description", "").lower()
                ):
                    continue
                if show_only_modified and fname not in overrides_table:
                    continue
                filtered.append(fname)

            if not filtered:
                st.caption("_Aucun champ ne correspond aux filtres._")
                continue

            for field_name in filtered:
                _render_field_priority_card(
                    table_slug,
                    field_name,
                    fields_spec[field_name],
                    overrides_table,
                    uploaded_slugs,
                )


def _render_field_priority_card(
    table_slug: str,
    field_name: str,
    field_spec: dict,
    overrides_table: dict,
    uploaded_slugs: set[str],
) -> None:
    """One card per field with reorderable priority list.

    When no override is active, the displayed priorities come from the YAML
    and preserve ties (ex-æquo). Once the user reorders anything, the override
    is recorded and priorities flatten to 1..N.
    """
    default_order = rules_io.default_priority_order(field_spec)  # [(slug, label, prio)]
    default_slugs = [s for s, _, _ in default_order]

    current_detailed = rules_io.resolve_current_order(
        field_spec, overrides_table.get(field_name)
    )  # [(slug, label, effective_priority)]
    current_slugs = [s for s, _, _ in current_detailed]

    # An override that equals the default order is noise — clean it up.
    if field_name in overrides_table and current_slugs == default_slugs:
        overrides_table.pop(field_name, None)
    is_modified = field_name in overrides_table

    mandatory = field_spec.get("mandatory", False)
    description = field_spec.get("description", "") or ""

    # Count ties per effective priority (for the ex-æquo badge).
    # Only meaningful when NOT modified — an override always flattens ties.
    prio_counts: dict[int, int] = {}
    for _, _, p in current_detailed:
        prio_counts[p] = prio_counts.get(p, 0) + 1

    # Card header badges
    badges: list[str] = []
    if mandatory:
        badges.append("🔒 Obligatoire")
    if is_modified:
        badges.append("✎ Modifié cette session")
    header_badges = ("  ·  " + "  ·  ".join(badges)) if badges else ""

    expander_title = f"**{field_name}** — {description}{header_badges}"
    with st.expander(expander_title, expanded=is_modified):
        # "Before / after" helper when modified
        if is_modified:
            labels_by_slug = {s: lbl for s, lbl, _ in default_order}
            c_before, c_after = st.columns(2)
            with c_before:
                st.caption("Priorités par défaut (ex-æquo respectés)")
                st.markdown(
                    " · ".join(
                        f"`{labels_by_slug.get(s, s)}` _(p{p})_"
                        for s, _, p in default_order
                    )
                )
            with c_after:
                st.caption("Priorités pour cet onboarding (ordre strict)")
                st.markdown(
                    " · ".join(
                        f"**`{lbl}`** _(p{p})_"
                        for _, lbl, p in current_detailed
                    )
                )
            st.markdown("")

        # Priority list with up/down arrows. Prio number shown comes straight
        # from the YAML when not overridden (ties visible), or flat 1..N when
        # overridden.
        for pos, (slug, label, prio) in enumerate(current_detailed):
            uploaded_mark = "🟢" if slug in uploaded_slugs else "⚪"
            is_tied = prio_counts.get(prio, 0) > 1 and not is_modified
            tie_badge = (
                "<span style='color:#b26a00;font-size:0.8em;"
                "margin-left:0.5em;font-style:italic'>· ex-æquo</span>"
                if is_tied
                else ""
            )

            row_cols = st.columns([1, 6, 1, 1])
            with row_cols[0]:
                st.markdown(f"### {prio}")
            with row_cols[1]:
                st.markdown(
                    f"{uploaded_mark} **{label}**{tie_badge}  \n"
                    f"<span style='color:#888;font-size:0.85em'>`{slug}`</span>",
                    unsafe_allow_html=True,
                )
            with row_cols[2]:
                if pos > 0:
                    if st.button(
                        "⬆",
                        key=f"up_{table_slug}_{field_name}_{slug}",
                        help="Monter d'un cran (casse les ex-æquo)",
                        use_container_width=True,
                    ):
                        new_order = list(current_slugs)
                        new_order[pos], new_order[pos - 1] = (
                            new_order[pos - 1],
                            new_order[pos],
                        )
                        overrides_table[field_name] = new_order
                        st.rerun()
            with row_cols[3]:
                if pos < len(current_detailed) - 1:
                    if st.button(
                        "⬇",
                        key=f"down_{table_slug}_{field_name}_{slug}",
                        help="Descendre d'un cran (casse les ex-æquo)",
                        use_container_width=True,
                    ):
                        new_order = list(current_slugs)
                        new_order[pos], new_order[pos + 1] = (
                            new_order[pos + 1],
                            new_order[pos],
                        )
                        overrides_table[field_name] = new_order
                        st.rerun()

        # Footer
        ft1, ft2 = st.columns([4, 1])
        with ft1:
            nb_uploaded = sum(1 for s in current_slugs if s in uploaded_slugs)
            mode_str = (
                "ordre strict appliqué"
                if is_modified
                else "priorités par défaut (ex-æquo respectés)"
            )
            st.caption(
                f"🟢 {nb_uploaded}/{len(current_slugs)} source(s) chargée(s) · {mode_str}."
            )
        with ft2:
            if is_modified:
                if st.button(
                    "↺ Réinitialiser",
                    key=f"reset_{table_slug}_{field_name}",
                    help="Revenir à l'ordre par défaut pour ce champ",
                    use_container_width=True,
                ):
                    overrides_table.pop(field_name, None)
                    st.rerun()


if st.session_state.mode == "home":
    render_home_page()

elif st.session_state.mode == "engine":
    render_engine_page()

elif st.session_state.mode == "rules":
    render_rules_page()

elif st.session_state.mode == "normal":
    render_normalization_page()

# ========== Classic mode: horizontal stepper above step content ==========
elif st.session_state.mode == "classic":
    _render_classic_stepper()

# The step branches below only fire when mode == "classic" (other modes have
# already returned above). Each uses `mode == classic` as a safety guard so a
# stale session_state.step never leaks into home/engine/rules pages.
_IS_CLASSIC = st.session_state.mode == "classic"

# ========== Step 1: Upload ==========
if _IS_CLASSIC and st.session_state.step == 1:
    st.header("1. Upload des fichiers")
    st.markdown(
        "Dépose ici **tous les fichiers** reçus pour ce client : fichiers internes, "
        "exports loueurs (Ayvens, Arval, etc.), fichier driver du client, et idéalement "
        "l'export API Plaques si tu l'as déjà fait."
    )
    uploaded = st.file_uploader(
        "Fichiers CSV / XLSX",
        type=["csv", "xlsx", "xls", "xlsm"],
        accept_multiple_files=True,
    )

    if uploaded:
        progress = st.progress(0.0, text="Analyse des fichiers...")
        new_sources: dict[str, SourceFile] = {}
        for i, up in enumerate(uploaded):
            try:
                pairs = load_tabular(up)
            except Exception as e:
                st.error(f"Impossible de lire {up.name}: {e}")
                continue
            for sheet_name, df in pairs:
                if df.empty or df.shape[1] < 2:
                    continue
                key = f"{up.name}::{sheet_name}" if sheet_name else up.name
                detected = detect(up.name, df, sheet_name=sheet_name or None)
                # Default target schema = first feed.
                default_target = detected.feeds[0] if detected.feeds else None
                new_sources[key] = SourceFile(
                    key=key,
                    filename=up.name,
                    sheet_name=sheet_name,
                    df_raw=df,
                    detected=detected,
                    target_schema=default_target,
                )
            progress.progress((i + 1) / len(uploaded), text=f"Analyse: {up.name}")
        st.session_state.sources = new_sources
        progress.empty()
        st.success(f"{len(new_sources)} fichier(s)/onglet(s) chargé(s). Passe à l'étape 2.")

    if st.session_state.sources:
        st.markdown("#### Résumé")
        rows = []
        for sf in st.session_state.sources.values():
            rows.append({
                "Fichier": sf.filename + (f" [{sf.sheet_name}]" if sf.sheet_name else ""),
                "Type détecté": label_for(sf.detected.source_type),
                "Alimente": ", ".join(sf.detected.feeds) or "—",
                "Lignes": len(sf.df),
            })
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


# ========== Step 2: Detection review ==========
elif _IS_CLASSIC and st.session_state.step == 2:
    st.header("2. Détection & choix du schéma cible")
    st.caption(
        "Vérifie le type détecté pour chaque fichier. "
        "Un fichier peut alimenter plusieurs schémas (ex. Ayvens Etat de parc → véhicules + contrats) : "
        "coche ceux qui t'intéressent."
    )

    if not st.session_state.sources:
        st.info("Rien à traiter. Remonte à l'étape 1.")
    else:
        for key, sf in st.session_state.sources.items():
            with st.expander(f"📄 {sf.filename}" + (f" [{sf.sheet_name}]" if sf.sheet_name else ""), expanded=True):
                c1, c2, c3 = st.columns([2, 2, 1])
                with c1:
                    st.markdown(f"**Détection**: {label_for(sf.detected.source_type)}")
                    st.caption(sf.detected.reason)
                with c2:
                    all_schemas = list(SCHEMAS.keys())
                    current = sf.target_schema or (all_schemas[0] if not sf.detected.feeds else sf.detected.feeds[0])
                    sf.target_schema = st.selectbox(
                        "Schéma Revio cible",
                        options=all_schemas + ["(ignorer ce fichier)"],
                        index=(all_schemas.index(current) if current in all_schemas else len(all_schemas)),
                        key=f"target_{key}",
                    )
                    if sf.target_schema == "(ignorer ce fichier)":
                        sf.selected = False
                        sf.target_schema = None
                    else:
                        sf.selected = True
                with c3:
                    st.metric("Lignes", len(sf.df))
                st.dataframe(sf.df.head(5), use_container_width=True)


# ========== Step 3: Column mapping ==========
elif _IS_CLASSIC and st.session_state.step == 3:
    st.header("3. Mapping des colonnes")
    st.caption(
        "Pour chaque fichier, je propose un mapping automatique. Révise-le si besoin. "
        "Les champs non mappés resteront vides dans la sortie."
    )
    if not st.session_state.sources:
        st.info("Rien à traiter.")
    for key, sf in st.session_state.sources.items():
        if not sf.selected or not sf.target_schema:
            continue
        with st.expander(
            f"🔗 {sf.filename} → {sf.target_schema.upper()}"
            + (f" [{sf.sheet_name}]" if sf.sheet_name else ""),
            expanded=False,
        ):
            c1, c2 = st.columns([1, 2])
            with c1:
                if st.button("🤖 Proposer un mapping (IA)", key=f"llm_{key}"):
                    with st.spinner("Claude analyse les colonnes..."):
                        result = propose_mapping(
                            sf.df, sf.target_schema, st.session_state.user_instructions
                        )
                    if "_error" in result:
                        st.error(result["_error"])
                    else:
                        proposed = result.get("mapping", {})
                        sf.mapping.update(proposed)
                        notes = result.get("_notes", [])
                        debug = result.get("_debug", {})
                        nb_mapped = sum(1 for v in proposed.values() if v)
                        if nb_mapped == 0:
                            st.warning(
                                "L'IA n'a proposé aucun mapping exploitable. "
                                "Voir le détail ci-dessous, et/ou mappe manuellement."
                            )
                        else:
                            st.success(
                                f"Mapping proposé : {nb_mapped} champ(s) rempli(s). "
                                "Révise-le ci-dessous."
                            )
                        if notes:
                            st.info("Notes de l'IA: " + " / ".join(str(n) for n in notes))
                        if debug:
                            with st.expander("🔍 Debug IA (pour diagnostic)"):
                                st.json({"mapping_proposé": proposed, **debug})
            with c2:
                st.caption(
                    "Colonnes détectées dans le fichier : "
                    + ", ".join(f"`{c}`" for c in list(sf.df.columns)[:10])
                    + (" ..." if len(sf.df.columns) > 10 else "")
                )
            st.markdown("---")
            target_fields = [f.name for f in SCHEMAS[sf.target_schema] if f.name]
            source_cols = [""] + list(sf.df.columns.astype(str))
            # Build a per-field selectbox.
            for tf in target_fields:
                current = sf.mapping.get(tf, "") or ""
                if current not in source_cols:
                    current = ""
                sf.mapping[tf] = st.selectbox(
                    tf,
                    options=source_cols,
                    index=source_cols.index(current),
                    key=f"map_{key}_{tf}",
                )


# ========== Step 4: Fleet splitting ==========
elif _IS_CLASSIC and st.session_state.step == 4:
    st.header("4. Découpage des flottes")
    st.caption(
        "L'outil détecte les agences présentes dans tes fichiers. "
        "Associe chacune à une flotte (ex. 1 flotte par agence, ou regroupements)."
    )

    # Collect all agencies from any source file that has an agency-like column.
    agencies: set[str] = set()
    AGENCY_CANDIDATES = ["agence", "agency", "structure", "centre", "site", "établissement", "etablissement",
                         "companyAnalyticalCode", "Structure 1"]
    for sf in st.session_state.sources.values():
        if not sf.selected:
            continue
        df = sf.df
        for col in df.columns:
            if any(cand.lower() in str(col).lower() for cand in AGENCY_CANDIDATES):
                agencies.update(str(v).strip() for v in df[col].dropna().unique() if str(v).strip())
    agencies = sorted(a for a in agencies if a and a.lower() not in {"nan", "none", "null"})

    if not agencies:
        st.info("Aucune agence détectée - tout ira dans une flotte 'default'.")
    else:
        c1, c2 = st.columns([1, 3])
        with c1:
            if st.button("🔁 Mode rapide : 1 agence = 1 flotte"):
                st.session_state.fleet_mapping = {a: a for a in agencies}
        st.markdown(f"**{len(agencies)} agence(s) détectée(s)**")
        current_fleets: set[str] = set(st.session_state.fleet_mapping.values()) or {"default"}
        for agency in agencies:
            current = st.session_state.fleet_mapping.get(agency, agency)
            new_val = st.text_input(
                f"Agence `{agency}` → Flotte",
                value=current,
                key=f"fleet_{agency}",
            )
            st.session_state.fleet_mapping[agency] = new_val


# ========== Step 5: Outputs & errors ==========
elif _IS_CLASSIC and st.session_state.step == 5:
    st.header("5. Génération des fichiers Revio")
    if not st.session_state.sources:
        st.info("Rien à traiter.")
        st.stop()

    # Priority for conflict resolution. Lower index = higher priority.
    priority_order = [
        "api_plaques",
        "ayvens_etat_parc",
        "arval_uat",
        "ayvens_aen",
        "arval_aen",
        "client_driver",
        "client_vehicle",
    ]

    with st.spinner("Construction des fichiers de sortie..."):
        sources_list = list(st.session_state.sources.values())
        outputs = merge_per_schema(sources_list, priority_order)
        # Split by fleet (driver files have companyAnalyticalCode filled naturally,
        # but we don't always have it on vehicle/contract outputs - for the v0 we
        # keep them together and rely on join via plate later).
        outputs_by_fleet = split_by_fleet(outputs, st.session_state.fleet_mapping)
        issues = validate(outputs)

    st.subheader("📊 Aperçu")
    total_rows = sum(len(df) for fleet in outputs_by_fleet.values() for df in fleet.values())
    st.metric("Lignes générées (toutes flottes × schémas)", total_rows)

    for fleet, schemas in outputs_by_fleet.items():
        with st.expander(f"🏢 Flotte: {fleet}", expanded=len(outputs_by_fleet) <= 2):
            for schema_name, df in schemas.items():
                st.markdown(f"**{schema_name}** ({len(df)} lignes)")
                st.dataframe(df.head(20), use_container_width=True, hide_index=True)

    if issues:
        st.subheader(f"⚠️ {len(issues)} problème(s) détecté(s)")
        issue_df = pd.DataFrame([
            {
                "schema": i.schema,
                "plaque": i.plate,
                "champ": i.field,
                "niveau": i.level,
                "message": i.message,
            }
            for i in issues
        ])
        st.dataframe(issue_df, use_container_width=True, hide_index=True)
    else:
        st.success("Aucune erreur bloquante détectée ✅")

    # Build the zip.
    client_name = st.session_state.client_name or "client"
    zip_bytes = build_zip(outputs_by_fleet, issues, client_name=client_name)
    st.download_button(
        "⬇️ Télécharger l'archive (CSV Revio + rapport)",
        data=zip_bytes,
        file_name=f"revio_import_{client_name}.zip",
        mime="application/zip",
        type="primary",
        use_container_width=True,
    )
