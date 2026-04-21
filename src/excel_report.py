"""Build a 3-sheet Excel report from an EngineResult.

Sheets:
1. `source` — plates (rows) × output fields (cols). Cell = display name of the
   source that won for that cell (e.g. "SIV", "Ayvens - État de parc").
   Empty cell means no source produced a value (the output cell is empty too).
2. `anomalies` — same shape. Cell is filled ONLY when 2+ sources produced
   DIFFERENT values for the same (plate, field). Format:
       [gardé] SIV=105
       vs Ayvens - État de parc=108
   The winner is prefixed with [gardé].
3. `plaques_orphelines` — plates found in lessor files but absent from the
   client file, with the same columns as the main output + a `sources_found`
   column listing the lessor slugs where each orphan was found.

Public API:
    build_report_xlsx(engine_result, client_name: str) -> bytes
"""

from __future__ import annotations

from io import BytesIO
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Hardcoded display names for source slugs. Takes precedence over YAML
# `source_label` so that api_plaques reads "SIV" (the value every fleet
# manager knows) rather than "API Plaques" (internal nickname).
SOURCE_DISPLAY_OVERRIDES: dict[str, str] = {
    "api_plaques": "SIV",
    "__default__": "défaut",
    "__engine__": "moteur",
}


def _build_source_display_map(rules_yaml: Optional[dict]) -> dict[str, str]:
    """Return {slug: display_name} from the YAML + hardcoded overrides."""
    out: dict[str, str] = {}
    if rules_yaml:
        for _, spec in (rules_yaml.get("fields") or {}).items():
            for rule in spec.get("rules", []):
                slug = rule.get("source")
                label = rule.get("source_label")
                if slug and label and slug not in out:
                    out[slug] = label
    # Hardcoded overrides win
    out.update(SOURCE_DISPLAY_OVERRIDES)
    return out


def _display(slug: str, display_map: dict[str, str]) -> str:
    return display_map.get(slug, slug)


def _fmt_val(v: Any) -> str:
    """Format a value for inclusion in the anomalies cell."""
    if v is None:
        return "∅"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _autosize_columns(ws, max_width: int = 60) -> None:
    """Rough auto-size of column widths based on content length."""
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            s = str(cell.value)
            longest_line = max((len(line) for line in s.split("\n")), default=0)
            if longest_line > max_len:
                max_len = longest_line
        width = min(max(max_len + 2, 10), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------- Public API ----------


def build_report_xlsx(engine_result, client_name: str = "client") -> bytes:
    """Build a 3-sheet Excel report from an EngineResult. Returns bytes."""
    display_map = _build_source_display_map(engine_result.rules_yaml)

    df = engine_result.df
    source_by_cell = engine_result.source_by_cell or {}
    conflicts_by_cell = engine_result.conflicts_by_cell or {}
    orphan_df = engine_result.orphan_df

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    plate_font = Font(bold=True)
    plate_fill = PatternFill("solid", fgColor="D9E1F2")
    conflict_fill = PatternFill("solid", fgColor="FFE699")
    wrap_align = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="center")

    # ===== Sheet 1: source =====
    ws_src = wb.create_sheet("source")
    headers = ["plaque"] + list(df.columns)
    ws_src.append(headers)
    for c in ws_src[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align

    for plate_key in df.index:
        plate_display = df.at[plate_key, "registrationPlate"] if "registrationPlate" in df.columns else plate_key
        if plate_display is None or (isinstance(plate_display, float) and plate_display != plate_display):
            plate_display = plate_key
        row = [plate_display]
        for field_name in df.columns:
            src = source_by_cell.get((str(plate_key), field_name))
            row.append(_display(src, display_map) if src else "")
        ws_src.append(row)
        ws_src.cell(row=ws_src.max_row, column=1).font = plate_font
        ws_src.cell(row=ws_src.max_row, column=1).fill = plate_fill

    ws_src.freeze_panes = "B2"
    _autosize_columns(ws_src, max_width=40)

    # ===== Sheet 2: anomalies =====
    ws_ano = wb.create_sheet("anomalies")
    ws_ano.append(headers)
    for c in ws_ano[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align

    anomaly_count = 0
    for plate_key in df.index:
        plate_display = df.at[plate_key, "registrationPlate"] if "registrationPlate" in df.columns else plate_key
        if plate_display is None or (isinstance(plate_display, float) and plate_display != plate_display):
            plate_display = plate_key
        row = [plate_display]
        row_has_conflict = False
        for field_name in df.columns:
            conflicts = conflicts_by_cell.get((str(plate_key), field_name))
            if not conflicts:
                row.append("")
                continue
            winner_src, winner_val = conflicts[0]
            lines = [f"[gardé] {_display(winner_src, display_map)}={_fmt_val(winner_val)}"]
            for src, val in conflicts[1:]:
                lines.append(f"vs {_display(src, display_map)}={_fmt_val(val)}")
            row.append("\n".join(lines))
            row_has_conflict = True
            anomaly_count += 1
        ws_ano.append(row)
        row_idx = ws_ano.max_row
        ws_ano.cell(row=row_idx, column=1).font = plate_font
        ws_ano.cell(row=row_idx, column=1).fill = plate_fill
        if row_has_conflict:
            for col_idx, field_name in enumerate(df.columns, start=2):
                val = ws_ano.cell(row=row_idx, column=col_idx).value
                if val:
                    ws_ano.cell(row=row_idx, column=col_idx).alignment = wrap_align
                    ws_ano.cell(row=row_idx, column=col_idx).fill = conflict_fill

    ws_ano.freeze_panes = "B2"
    _autosize_columns(ws_ano, max_width=50)

    if anomaly_count == 0:
        ws_ano.insert_rows(2)
        ws_ano.cell(row=2, column=1).value = (
            "Aucune anomalie détectée : toutes les sources sont cohérentes entre elles."
        )
        ws_ano.cell(row=2, column=1).font = Font(italic=True, color="666666")
        ws_ano.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    # ===== Sheet 3: plaques_orphelines =====
    ws_orp = wb.create_sheet("plaques_orphelines")
    if orphan_df is None or orphan_df.empty:
        ws_orp.cell(row=1, column=1).value = (
            "Aucune plaque orpheline : toutes les plaques des fichiers loueurs sont "
            "présentes dans le fichier client."
        )
        ws_orp.cell(row=1, column=1).font = Font(italic=True, color="666666")
        ws_orp.column_dimensions["A"].width = 100
    else:
        orphan_headers = ["plaque"] + list(orphan_df.columns)
        ws_orp.append(orphan_headers)
        for c in ws_orp[1]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = center_align

        for plate_key in orphan_df.index:
            plate_display = plate_key
            if "registrationPlate" in orphan_df.columns:
                rp = orphan_df.at[plate_key, "registrationPlate"]
                if rp is not None and not (isinstance(rp, float) and rp != rp):
                    plate_display = rp
            row = [plate_display]
            for field_name in orphan_df.columns:
                v = orphan_df.at[plate_key, field_name]
                if v is None or (isinstance(v, float) and v != v):
                    row.append("")
                else:
                    row.append(v)
            ws_orp.append(row)
            ws_orp.cell(row=ws_orp.max_row, column=1).font = plate_font
            ws_orp.cell(row=ws_orp.max_row, column=1).fill = plate_fill

        ws_orp.freeze_panes = "B2"
        _autosize_columns(ws_orp, max_width=40)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
