"""Tests for src/zip_writer.py.

We check the zip structure (filenames, presence/absence of per-fleet
files when segmentation is active vs not), the CSV contents (Revio
header order + empty-placeholder column preserved), and the master
Excel tab plan.
"""

from __future__ import annotations

import io
import unittest
import zipfile

import pandas as pd
from openpyxl import load_workbook

from src import fleet_segmentation as fs
from src import zip_writer as zw
from src.schemas import header_for


def _sample_vehicle_df() -> pd.DataFrame:
    """A minimal vehicle output df indexed by plate_for_matching."""
    return pd.DataFrame(
        {
            "registrationPlate": ["AB-123-CD", "EF-456-GH", "IJ-789-KL"],
            "usage": ["private", "utility", "service"],
            "motorisation": ["diesel", "gas", "electric"],
            "registrationIssueCountryCode": ["FR", "FR", "FR"],
        },
        index=["AB123CD", "EF456GH", "IJ789KL"],
    )


class TestRevioCsv(unittest.TestCase):
    def test_header_order_matches_schema(self):
        df = _sample_vehicle_df()
        raw = zw.revio_csv_bytes(df, "vehicle").decode("utf-8-sig")
        header = raw.splitlines()[0].split(",")
        self.assertEqual(header, header_for("vehicle"))

    def test_preserves_empty_placeholder_column(self):
        """Vehicle schema has an empty-named column between co2gKm and
        registrationIssueCountryCode. The CSV should keep it as an empty
        header slot (→ two consecutive commas in the header line)."""
        df = _sample_vehicle_df()
        raw = zw.revio_csv_bytes(df, "vehicle").decode("utf-8-sig")
        header_line = raw.splitlines()[0]
        self.assertIn(",,", header_line)

    def test_missing_fields_produce_empty_cells(self):
        df = _sample_vehicle_df()
        raw = zw.revio_csv_bytes(df, "vehicle").decode("utf-8-sig")
        # There are 3 data rows.
        self.assertEqual(len(raw.splitlines()), 4)


class TestMasterXlsx(unittest.TestCase):
    def test_no_fleet_mapping_single_tab(self):
        df = _sample_vehicle_df()
        blob = zw.build_master_xlsx(
            client_name="ACME",
            vehicle_df=df,
            vehicle_fleet_mapping=None,
        )
        wb = load_workbook(io.BytesIO(blob))
        # Expect: Sommaire + vehicles — tous
        self.assertEqual(wb.sheetnames[0], "Sommaire")
        self.assertIn("vehicles — tous", wb.sheetnames)
        self.assertEqual(len(wb.sheetnames), 2)

    def test_with_fleet_mapping_one_tab_per_fleet_plus_tous(self):
        df = _sample_vehicle_df()
        mapping = fs.FleetMapping(
            source_file_key="x",
            source_column="Agence",
            raw_to_fleet={"bdx": "Bordeaux", "paris": "Paris"},
            plate_to_fleet={"AB123CD": "Bordeaux", "EF456GH": "Paris"},
        )
        blob = zw.build_master_xlsx(
            client_name="ACME",
            vehicle_df=df,
            vehicle_fleet_mapping=mapping,
        )
        wb = load_workbook(io.BytesIO(blob))
        # Sommaire + tous + Bordeaux + Paris + non rattaché (IJ789KL unassigned)
        self.assertIn("vehicles — tous", wb.sheetnames)
        self.assertIn("vehicles — Bordeaux", wb.sheetnames)
        self.assertIn("vehicles — Paris", wb.sheetnames)
        self.assertTrue(
            any("non rattach" in n.lower() for n in wb.sheetnames),
            f"Missing non-rattaché sheet in {wb.sheetnames}",
        )

    def test_empty_df_does_not_crash(self):
        blob = zw.build_master_xlsx(
            client_name="X",
            vehicle_df=pd.DataFrame(),
            vehicle_fleet_mapping=None,
        )
        wb = load_workbook(io.BytesIO(blob))
        self.assertIn("Sommaire", wb.sheetnames)


class TestBuildOutputZip(unittest.TestCase):
    def test_zip_structure_no_segmentation(self):
        df = _sample_vehicle_df()
        blob, filename = zw.build_output_zip(
            client_name="ACME SA",
            vehicle_df=df,
            vehicle_fleet_mapping=None,
        )
        self.assertTrue(filename.startswith("revio_import_acme-sa_"))
        self.assertTrue(filename.endswith(".zip"))

        with zipfile.ZipFile(io.BytesIO(blob)) as zf:
            names = zf.namelist()
        # Every path is prefixed with the root folder.
        root = filename[:-4]  # strip .zip
        self.assertIn(f"{root}/onboarding_complet.xlsx", names)
        self.assertIn(f"{root}/vehicles/vehicles_tous.csv", names)
        # No per-fleet files when no mapping.
        per_fleet = [n for n in names
                     if n.startswith(f"{root}/vehicles/vehicles_")
                     and not n.endswith("_tous.csv")]
        self.assertEqual(per_fleet, [])

    def test_zip_structure_with_segmentation(self):
        df = _sample_vehicle_df()
        mapping = fs.FleetMapping(
            source_file_key="x",
            source_column="Agence",
            raw_to_fleet={"bdx": "Bordeaux", "paris": "Paris"},
            plate_to_fleet={"AB123CD": "Bordeaux", "EF456GH": "Paris"},
        )
        blob, filename = zw.build_output_zip(
            client_name="ACME",
            vehicle_df=df,
            vehicle_fleet_mapping=mapping,
        )
        with zipfile.ZipFile(io.BytesIO(blob)) as zf:
            names = zf.namelist()
        root = filename[:-4]
        self.assertIn(f"{root}/vehicles/vehicles_tous.csv", names)
        self.assertIn(f"{root}/vehicles/vehicles_bordeaux.csv", names)
        self.assertIn(f"{root}/vehicles/vehicles_paris.csv", names)
        # IJ789KL had no fleet → non-rattaché CSV expected.
        self.assertTrue(
            any("non-rattache" in n for n in names),
            f"Missing non-rattaché CSV in {names}",
        )

    def test_zip_includes_report_when_provided(self):
        df = _sample_vehicle_df()
        blob, filename = zw.build_output_zip(
            client_name="ACME",
            vehicle_df=df,
            report_xlsx_bytes=b"fake xlsx bytes",
        )
        with zipfile.ZipFile(io.BytesIO(blob)) as zf:
            names = zf.namelist()
        root = filename[:-4]
        self.assertIn(f"{root}/rapport.xlsx", names)

    def test_csv_content_has_bom(self):
        df = _sample_vehicle_df()
        blob, filename = zw.build_output_zip(
            client_name="ACME",
            vehicle_df=df,
        )
        root = filename[:-4]
        with zipfile.ZipFile(io.BytesIO(blob)) as zf:
            csv_bytes = zf.read(f"{root}/vehicles/vehicles_tous.csv")
        # UTF-8 BOM
        self.assertTrue(csv_bytes.startswith(b"\xef\xbb\xbf"))


if __name__ == "__main__":
    unittest.main()
