"""Build a multi-sheet Excel audit report from an engine result.

Two entry points :

- :func:`build_report_xlsx` — Vehicle (3 sheets : ``source`` / ``anomalies``
  / ``plaques_orphelines``). One leading column "plaque" per row.

- :func:`build_contract_report_xlsx` — Contract (4 sheets : ``source`` /
  ``anomalies`` / ``plaques_orphelines`` / ``issues``). Two leading
  columns ("plaque" + "numéro") per row, since a contract is identified
  by a (plate, contract number) pair even though the engine keys by plate
  alone since Jalon 4.2.6. The 4th ``issues`` sheet keeps the long-form
  global warnings that don't have a target cell (e.g. "135 plates with
  multiple rows in ayvens_facture_pdf").

Both reports share the same internal grid builders so the visual format
is identical : ``source`` is a wide grid where every cell shows the slug
that won that field, ``anomalies`` mirrors that grid but only the cells
where 2+ sources produced different values get filled with
``[gardé] X=val`` + ``vs Y=val`` (winner first), highlighted in yellow.

Public API :
    build_report_xlsx(engine_result, client_name: str = "client") -> bytes
    build_contract_report_xlsx(engine_result, client_name: str, issues=...) -> bytes
"""

from __future__ import annotations

from io import BytesIO
from typing import Any, Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Hardcoded display names for source slugs. Takes precedence over YAML
# `source_label` so that api_plaques reads "SIV" (the value every fleet
# manager knows) rather than "API Plaques" (internal nickname).
SOURCE_DISPLAY_OVERRIDES: dict[str, str] = {
    "api_plaques": "SIV",
    "api_plaque": "SIV",  # singular form used by some lineage records
    "__default__": "défaut",
    "__engine__": "moteur",
    "rule_engine": "moteur",
    "derived": "dérivé",
}


def _build_source_display_map(rules_yaml: Optional[dict]) -> dict[str, str]:
    """Return {slug: display_name} from the YAML + hardcoded overrides."""
    out: dict[str, str] = {}
    if rules_yaml:
        for _, spec in (rules_yaml.get("fields") or {}).items():
            for rule in spec.get("rules", []):
                slug = rule.get("source")
                label = rule.get("source_label")
                if slug and label and slug not in out:
                    out[slug] = label
    # Hardcoded overrides win
    out.update(SOURCE_DISPLAY_OVERRIDES)
    return out


def _display(slug: str, display_map: dict[str, str]) -> str:
    return display_map.get(slug, slug)


def _fmt_val(v: Any) -> str:
    """Format a value for inclusion in the anomalies cell."""
    if v is None:
        return "∅"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _autosize_columns(ws, max_width: int = 60) -> None:
    """Rough auto-size of column widths based on content length."""
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            s = str(cell.value)
            longest_line = max((len(line) for line in s.split("\n")), default=0)
            if longest_line > max_len:
                max_len = longest_line
        width = min(max(max_len + 2, 10), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# =============================================================================
# Common styling
# =============================================================================

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
KEY_FONT = Font(bold=True)
KEY_FILL = PatternFill("solid", fgColor="D9E1F2")
CONFLICT_FILL = PatternFill("solid", fgColor="FFE699")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


def _safe_at(df, key, col):
    """Return ``df.at[key, col]`` defensively (NaN/None → empty string).

    Avoids the AttributeError class of bugs you get when ``key`` happens to
    not be in the index of the orphan DF (mixed indexes are rare but happen).
    """
    if df is None or col not in df.columns:
        return ""
    if key not in df.index:
        return ""
    v = df.at[key, col]
    if v is None:
        return ""
    if isinstance(v, float) and v != v:  # NaN
        return ""
    return v


# =============================================================================
# Generic grid sheet builders
# =============================================================================

def _row_leading_cells(df, key, key_columns: list[tuple[str, str]]) -> list:
    """Compute the 1-or-N leading display cells for a row.

    ``key_columns`` is a list of ``(header_label, source_column)`` pairs.
    For each pair we look up ``df.at[key, source_column]``. The FIRST
    column falls back to the index ``key`` when the value is missing
    (we always want a row identifier visible in the leftmost cell).
    Subsequent columns leave an empty cell when the source value is
    missing — e.g. a contract row with no ``number`` shouldn't display
    the plate key in the « numéro » column.
    """
    out = []
    for idx, (_label, src_col) in enumerate(key_columns):
        v = None
        if src_col and src_col in df.columns and key in df.index:
            v = df.at[key, src_col]
        if v is None or (isinstance(v, float) and v != v):
            out.append(key if idx == 0 else "")
        else:
            out.append(v)
    return out


def _write_source_sheet(
    ws,
    df,
    *,
    source_by_cell: dict,
    display_map: dict,
    key_columns: list[tuple[str, str]],
    field_columns: list[str],
) -> None:
    """Render the wide « source » grid into ``ws``.

    Cell value = display name of the slug that produced the winning value
    for ``(key, field)``. Empty cell ⇒ no source produced anything (final
    output cell is empty too).
    """
    headers = [label for label, _ in key_columns] + list(field_columns)
    ws.append(headers)
    for c in ws[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER_ALIGN

    for key in df.index:
        row = _row_leading_cells(df, key, key_columns)
        for field_name in field_columns:
            src = source_by_cell.get((str(key), field_name))
            row.append(_display(src, display_map) if src else "")
        ws.append(row)
        for col_idx in range(1, len(key_columns) + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.font = KEY_FONT
            cell.fill = KEY_FILL

    freeze_col = get_column_letter(len(key_columns) + 1)
    ws.freeze_panes = f"{freeze_col}2"
    _autosize_columns(ws, max_width=40)


def _write_anomalies_sheet(
    ws,
    df,
    *,
    conflicts_by_cell: dict,
    display_map: dict,
    key_columns: list[tuple[str, str]],
    field_columns: list[str],
) -> int:
    """Render the « anomalies » grid into ``ws``. Returns conflict count.

    Cell is filled ONLY for ``(key, field)`` where 2+ sources produced
    DIFFERENT values. Format::

        [gardé] SIV=105
        vs Ayvens - État de parc=108

    The winning source is prefixed with ``[gardé]``. Conflict cells are
    highlighted in yellow with wrap-text alignment so multi-line values
    stay readable.
    """
    headers = [label for label, _ in key_columns] + list(field_columns)
    ws.append(headers)
    for c in ws[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER_ALIGN

    n_kc = len(key_columns)
    anomaly_count = 0
    for key in df.index:
        row = _row_leading_cells(df, key, key_columns)
        row_has_conflict = False
        for field_name in field_columns:
            conflicts = conflicts_by_cell.get((str(key), field_name))
            if not conflicts:
                row.append("")
                continue
            winner_src, winner_val = conflicts[0]
            lines = [f"[gardé] {_display(winner_src, display_map)}={_fmt_val(winner_val)}"]
            for src, val in conflicts[1:]:
                lines.append(f"vs {_display(src, display_map)}={_fmt_val(val)}")
            row.append("\n".join(lines))
            row_has_conflict = True
            anomaly_count += 1
        ws.append(row)
        row_idx = ws.max_row
        for col_idx in range(1, n_kc + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = KEY_FONT
            cell.fill = KEY_FILL
        if row_has_conflict:
            for col_idx, _field_name in enumerate(field_columns, start=n_kc + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    ws.cell(row=row_idx, column=col_idx).alignment = WRAP_ALIGN
                    ws.cell(row=row_idx, column=col_idx).fill = CONFLICT_FILL

    freeze_col = get_column_letter(n_kc + 1)
    ws.freeze_panes = f"{freeze_col}2"
    _autosize_columns(ws, max_width=50)

    if anomaly_count == 0:
        ws.insert_rows(2)
        ws.cell(row=2, column=1).value = (
            "Aucune anomalie détectée : toutes les sources sont cohérentes entre elles."
        )
        ws.cell(row=2, column=1).font = Font(italic=True, color="666666")
        ws.merge_cells(
            start_row=2, start_column=1,
            end_row=2, end_column=len(headers),
        )

    return anomaly_count


def _write_orphans_sheet(
    ws,
    orphan_df,
    *,
    key_columns: list[tuple[str, str]],
    empty_message: str = (
        "Aucune plaque orpheline : toutes les plaques des fichiers loueurs sont "
        "présentes dans le fichier client."
    ),
) -> None:
    """Render orphans (rows present in lessor sources but absent from
    client_file). Same leading columns as the main grid, then the raw
    columns of ``orphan_df`` (no source attribution)."""
    if orphan_df is None or orphan_df.empty:
        ws.cell(row=1, column=1).value = empty_message
        ws.cell(row=1, column=1).font = Font(italic=True, color="666666")
        ws.column_dimensions["A"].width = 100
        return

    headers = [label for label, _ in key_columns] + list(orphan_df.columns)
    ws.append(headers)
    for c in ws[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER_ALIGN

    n_kc = len(key_columns)
    for key in orphan_df.index:
        row = _row_leading_cells(orphan_df, key, key_columns)
        for field_name in orphan_df.columns:
            v = orphan_df.at[key, field_name]
            if v is None or (isinstance(v, float) and v != v):
                row.append("")
            else:
                row.append(v)
        ws.append(row)
        for col_idx in range(1, n_kc + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.font = KEY_FONT
            cell.fill = KEY_FILL

    freeze_col = get_column_letter(n_kc + 1)
    ws.freeze_panes = f"{freeze_col}2"
    _autosize_columns(ws, max_width=40)


def _write_issues_sheet(ws, issues: Optional[Iterable]) -> None:
    """Render the long-form ``issues`` sheet for the Contract report.

    ``issues`` is whatever the engine attached to ``result.issues`` —
    usually a list of dataclass instances with attributes ``plate`` /
    ``number`` / ``field`` / ``source`` / ``warning``. We accept any iterable
    of objects (or dicts) and pull the same five attributes via
    ``getattr`` / ``.get`` so this works without coupling to a specific
    type.
    """
    headers = ["plaque", "numéro", "field", "source", "avertissement"]
    ws.append(headers)
    for c in ws[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER_ALIGN

    n = 0
    for i in (issues or []):
        if isinstance(i, dict):
            row = [
                i.get("plate", ""),
                i.get("number", ""),
                i.get("field", ""),
                i.get("source", ""),
                i.get("warning", ""),
            ]
        else:
            row = [
                getattr(i, "plate", "") or "",
                getattr(i, "number", "") or "",
                getattr(i, "field", "") or "",
                getattr(i, "source", "") or "",
                getattr(i, "warning", "") or "",
            ]
        ws.append(row)
        n += 1
    if n == 0:
        ws.cell(row=2, column=1).value = "Aucun avertissement transverse."
        ws.cell(row=2, column=1).font = Font(italic=True, color="666666")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    _autosize_columns(ws, max_width=80)


# =============================================================================
# Public API — Vehicle
# =============================================================================

def build_report_xlsx(engine_result, client_name: str = "client") -> bytes:
    """Build a 3-sheet Excel report from a Vehicle EngineResult. Returns bytes.

    Sheets : ``source`` / ``anomalies`` / ``plaques_orphelines``. One leading
    column "plaque" per row (filled from ``df["registrationPlate"]``, fallback
    to the index key).
    """
    display_map = _build_source_display_map(engine_result.rules_yaml)

    df = engine_result.df
    source_by_cell = engine_result.source_by_cell or {}
    conflicts_by_cell = engine_result.conflicts_by_cell or {}
    orphan_df = engine_result.orphan_df

    wb = Workbook()
    wb.remove(wb.active)

    key_columns = [("plaque", "registrationPlate")]
    field_columns = list(df.columns)

    _write_source_sheet(
        wb.create_sheet("source"),
        df,
        source_by_cell=source_by_cell,
        display_map=display_map,
        key_columns=key_columns,
        field_columns=field_columns,
    )
    _write_anomalies_sheet(
        wb.create_sheet("anomalies"),
        df,
        conflicts_by_cell=conflicts_by_cell,
        display_map=display_map,
        key_columns=key_columns,
        field_columns=field_columns,
    )
    _write_orphans_sheet(
        wb.create_sheet("plaques_orphelines"),
        orphan_df,
        key_columns=key_columns,
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# Public API — Contract (Jalon 5.2.1)
# =============================================================================

def build_contract_report_xlsx(
    engine_result,
    *,
    client_name: str = "client",
    issues: Optional[Iterable] = None,
) -> bytes:
    """Build a 4-sheet Excel report from a ContractEngineResult.

    Sheets :

    1. ``source``  — wide grid : ``(plaque, numéro)`` × all 45 contract
       fields. Cell = slug that won that field. Empty cell ⇒ no source
       produced anything (final cell is empty too).
    2. ``anomalies`` — same grid shape ; only conflict cells filled with
       ``[gardé] X=val \\n vs Y=val``.
    3. ``plaques_orphelines`` — contracts in lessor sources but absent
       from ``client_file``.
    4. ``issues`` — long-form list of transverse warnings (e.g. "135 plates
       with multiple rows in ayvens_facture_pdf"). One row per warning,
       ``(plaque, numéro, field, source, message)`` columns.

    The 2-leading-column shape (``plaque`` + ``numéro``) is the only
    visible difference vs the Vehicle report — and necessary because a
    single plate can carry several contracts (active + previous, or
    inter-loueur).

    ``issues`` defaults to ``engine_result.issues`` when not provided ;
    pass an explicit iterable to override.
    """
    display_map = _build_source_display_map(getattr(engine_result, "rules_yaml", None))

    df = engine_result.df
    source_by_cell = engine_result.source_by_cell or {}
    conflicts_by_cell = engine_result.conflicts_by_cell or {}
    orphan_df = engine_result.orphan_df
    if issues is None:
        issues = getattr(engine_result, "issues", None) or []

    wb = Workbook()
    wb.remove(wb.active)

    # Contract DF has both ``plate`` (display, with hyphens) and ``number``.
    # Index is the canonical plate key (post-Jalon 4.2.6 — plate-only keys).
    key_columns = [("plaque", "plate"), ("numéro", "number")]
    field_columns = list(df.columns)

    _write_source_sheet(
        wb.create_sheet("source"),
        df,
        source_by_cell=source_by_cell,
        display_map=display_map,
        key_columns=key_columns,
        field_columns=field_columns,
    )
    _write_anomalies_sheet(
        wb.create_sheet("anomalies"),
        df,
        conflicts_by_cell=conflicts_by_cell,
        display_map=display_map,
        key_columns=key_columns,
        field_columns=field_columns,
    )
    _write_orphans_sheet(
        wb.create_sheet("plaques_orphelines"),
        orphan_df,
        key_columns=key_columns,
        empty_message=(
            "Aucun contrat orphelin : tous les contrats des fichiers loueurs sont "
            "rattachés à une plaque du fichier client."
        ),
    )
    _write_issues_sheet(
        wb.create_sheet("issues"),
        issues,
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
